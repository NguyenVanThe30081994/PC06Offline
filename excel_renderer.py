"""
excel_renderer.py
-----------------
Shared utility: renders an openpyxl worksheet as a faithful HTML <table>,
preserving merged cells (colspan/rowspan), cell styles, and optionally
filling in submitted data values.

Used by:
  - V1 Stats page  (render header rows + append unit data rows)
  - V2 Render page (render full template, turn input-marker cells into <input>)
"""

import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.fills import PatternFill
from markupsafe import Markup


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _safe_color(color_obj):
    """Return a 6-char hex string or None from an openpyxl Color object."""
    try:
        if not color_obj:
            return None
        # RGB type
        if color_obj.type == 'rgb' and color_obj.rgb:
            rgb = str(color_obj.rgb).upper()
            if len(rgb) == 8: return rgb[2:]
            return rgb
        # Theme type (we return a marker or None as we don't resolve full themes here)
        if color_obj.type == 'theme' and color_obj.theme is not None:
            return f"THEME_{color_obj.theme}"
    except Exception:
        pass
    return None


def is_input_cell(cell):
    """
    Standard logic to detect if a cell is an input marker.
    Matches the project's light blue color (RGB E0F2FE) or its Theme/Indexed equivalent.
    """
    try:
        fill = cell.fill
        if not fill or (fill.patternType != 'solid' and fill.patternType is not None):
            return False
        
        c = fill.start_color
        if not c: return False
        
        # 1. Match by RGB (most common)
        INPUT_MARKERS = [
            'FFE0F2FE', '00E0F2FE', 'E0F2FE',  # Pro-Blue
            'FFB4C6E7', 'B4C6E7',             # Light Blue Accent
            'FFD9E1F2', 'D9E1F2',             # Very Light Blue
            'B7DEE8', 'FFB7DEE8',             # Aqua Light
            'DAE3F3', 'FFDAE3F3'              # Default Input Shade
        ]
        
        # Check rgb property
        if hasattr(c, 'rgb') and c.rgb:
            rgb = str(c.rgb).upper()
            if rgb in INPUT_MARKERS or any(m in rgb for m in INPUT_MARKERS):
                return True
        
        # Check indexed/index property (sometimes used for legacy or specific palettes)
        if hasattr(c, 'index'):
            idx = str(c.index).upper()
            if idx in INPUT_MARKERS or any(m in idx for m in INPUT_MARKERS):
                return True
            # Common indexed colors for light blue (depends on palette, but 40, 41 are often blue-ish)
            if idx in ['40', '41', '42']:
                return True

        # 2. Match by Theme (Theme 4, 5, 8 are common blue accents)
        if c.type == 'theme' and c.theme in [1, 4, 5, 8]:
            # Theme 1 + Tint -0.15 is often a light blue-ish gray used for inputs
            # Theme 4, 5, 8 are standard Accents (Blue, Gold, Aqua)
            if c.theme == 1 and (c.tint and -0.2 < c.tint < 0):
                return True
            if c.theme in [4, 5, 8]:
                return True
            
    except Exception:
        pass
    return False


def _cell_css(cell):
    """Build inline CSS string from openpyxl cell formatting."""
    parts = []

    # Background
    try:
        fill = cell.fill
        if fill and isinstance(fill, PatternFill) and fill.patternType == 'solid':
            hex_c = _safe_color(fill.fgColor)
            if hex_c:
                parts.append(f'background-color:#{hex_c}')
    except Exception:
        pass

    # Font
    try:
        font = cell.font
        if font:
            if font.bold:
                parts.append('font-weight:bold')
            if font.italic:
                parts.append('font-style:italic')
            if font.size:
                parts.append(f'font-size:{font.size}pt')
            hex_c = _safe_color(font.color) if font.color else None
            if hex_c:
                parts.append(f'color:#{hex_c}')
    except Exception:
        pass

    # Alignment
    try:
        al = cell.alignment
        if al:
            if al.horizontal:
                parts.append(f'text-align:{al.horizontal}')
            if al.vertical:
                parts.append(f'vertical-align:{al.vertical}')
            if al.wrap_text:
                parts.append('white-space:pre-wrap')
    except Exception:
        pass

    return ';'.join(parts)


def _build_merge_lookup(ws):
    """
    Returns:
      spans   : dict (row, col) -> (rowspan, colspan)  for merge top-left cells
      shadows : set of (row, col) that are covered by a merge (must be skipped)
    """
    spans = {}
    shadows = set()
    for mr in ws.merged_cells.ranges:
        rs = mr.max_row - mr.min_row + 1
        cs = mr.max_col - mr.min_col + 1
        spans[(mr.min_row, mr.min_col)] = (rs, cs)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    shadows.add((r, c))
    return spans, shadows


def _col_widths_px(ws):
    """Return list of pixel widths (1-indexed → 0-indexed list)."""
    widths = []
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        w = ws.column_dimensions[letter].width or 8.43
        widths.append(max(int(w * 7), 45))
    return widths


def _row_height_px(ws, r):
    h = ws.row_dimensions[r].height or 15
    return max(int(h * 1.33), 20)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def render_range_to_html(ws, start_row, end_row,
                         input_marker_hex='FFE0F2FE',
                         existing_values=None,
                         editable=True,
                         min_col=1, max_col=None):
    """
    Render rows [start_row .. end_row] (1-indexed, inclusive) of `ws`
    as an HTML <tbody> fragment.

    Parameters
    ----------
    ws               : openpyxl Worksheet
    start_row        : int
    end_row          : int
    input_marker_hex : str  – background fill hex (8-char ARGB) that marks input cells
    existing_values  : dict – cell_coord -> value  e.g. {'B5': '42'}
    editable         : bool – if False, all cells render as text (for stats display)
    min_col          : int  – first column to render (1-indexed)
    max_col          : int  – last column to render (1-indexed), defaults to ws.max_column

    Returns
    -------
    dict with:
      'tbody_html'  : str  – the <tbody>…</tbody> HTML
      'input_keys'  : list of cell coordinates that are input cells
    """
    if existing_values is None:
        existing_values = {}
    if max_col is None:
        max_col = ws.max_column

    spans, shadows = _build_merge_lookup(ws)
    html = []
    input_keys = []

    for r in range(start_row, end_row + 1):
        if ws.row_dimensions[r].hidden:
            continue
        rh = _row_height_px(ws, r)
        html.append(f'<tr style="height:{rh}px">')

        for c in range(min_col, max_col + 1):
            if (r, c) in shadows:
                continue

            cell = ws.cell(row=r, column=c)
            rowspan, colspan = spans.get((r, c), (1, 1))
            css = _cell_css(cell)

            # Detect input cell
            is_input = is_input_cell(cell)

            coord = cell.coordinate
            rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
            cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
            base_style = (
                'padding:3px 6px;border:1px solid #d1d5db;'
                'overflow:hidden;box-sizing:border-box;'
            )
            if is_input:
                base_style += 'background-color:#e0f2fe;'

            full_css = f'{base_style}{css}'
            td_open = f'<td{rs_attr}{cs_attr} style="{full_css}">'

            if is_input:
                val = existing_values.get(coord, '')
                safe_val = str(val).replace('"', '&quot;')
                input_keys.append(coord)
                td_inner = (
                    f'<input type="text" class="grid-input" '
                    f'data-key="{coord}" data-coord="{coord}" '
                    f'value="{safe_val}" onchange="markDirty()" '
                    f'style="width:100%;height:100%;border:none;'
                    f'background:transparent;padding:2px;font-size:inherit;">'
                )
            else:
                raw_val = cell.value
                # For formula cells use openpyxl cached value (data_only mode)
                if isinstance(raw_val, str) and raw_val.startswith('='):
                    raw_val = ''
                display = '' if raw_val is None else str(raw_val)
                td_inner = display

            html.append(f'{td_open}{td_inner}</td>')

        html.append('</tr>')

    return {
        'tbody_html': '\n'.join(html),
        'input_keys': input_keys,
    }



def build_stats_table_html(file_blob, config, submissions):
    """
    For V1 Stats: render the Excel file's header section faithfully,
    then append one data row per unit below the headers.

    Parameters
    ----------
    file_blob   : bytes  – raw Excel file from ReportConfig.file_blob
    config      : ReportConfig instance
    submissions : list of dicts:
                  [{'unit': str, 'date': str, 'sender': str, 'values': {idx_str: val}}, ...]

    Returns
    -------
    Markup – complete <table>...</table> HTML safe string
    """
    if not file_blob:
        return Markup('<p class="text-muted">Không có file Excel gốc trong hệ thống.</p>')

    try:
        # Use data_only=True to get cached formula values
        wb = openpyxl.load_workbook(io.BytesIO(file_blob), data_only=True)
        ws = wb.active
    except Exception as e:
        return Markup(f'<p class="text-danger">Lỗi đọc file Excel: {e}</p>')

    header_start = config.header_start or 1
    header_rows = config.header_rows or 1
    header_end = header_start + header_rows - 1
    # ---------- Determine used range ----------
    from pc06_excel_engine import ExcelEngineV2
    regions = ExcelEngineV2._detect_active_regions(wb, ws)
    r_box = regions["report"]
    min_col, min_row, max_col, max_row = r_box[0], r_box[1], r_box[2], r_box[3]
    
    # Ensure header_start is not excluded if it's before min_row
    render_start_row = min(header_start, min_row)

    import json
    try:
        fields = json.loads(config.config_json or '[]')
    except Exception:
        fields = []

    # Map Unit Name -> Submission (Case-insensitive for matching)
    unit_map = {sub.get('unit'): sub for sub in submissions}
    unit_map_lower = {str(k).strip().lower(): v for k, v in unit_map.items() if k}
    unit_names_lower = sorted(list(unit_map_lower.keys()), key=len, reverse=True)

    spans, shadows = _build_merge_lookup(ws)
    col_widths = []
    for i in range(min_col, max_col + 1):
        letter = get_column_letter(i)
        w = ws.column_dimensions[letter].width or 8.43
        col_widths.append(max(int(w * 7), 45))

    # ---------- colgroup ----------
    col_parts = ['<colgroup>']
    for w in col_widths:
        col_parts.append(f'<col style="width:{w}px">')
    col_parts.append('</colgroup>')

    # ---------- Render ALL rows from render_start_row to max_row ----------
    rows_html = []
    for r in range(render_start_row, max_row + 1):
        if ws.row_dimensions[r].hidden:
            continue
        
        # 1. Identify if this row belongs to a unit (only search within min_col to max_col)
        row_content = ""
        for c in range(min_col, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val: row_content += str(val)
        
        row_content_lower = row_content.lower()
        matched_sub = None
        for name in unit_names_lower:
            if name and name in row_content_lower:
                matched_sub = unit_map_lower[name]
                break
        
        rh = _row_height_px(ws, r)
        rows_html.append(f'<tr style="height:{rh}px">')

        for c in range(min_col, max_col + 1):
            if (r, c) in shadows:
                continue

            cell = ws.cell(row=r, column=c)
            rowspan, colspan = spans.get((r, c), (1, 1))
            css = _cell_css(cell)
            
            # If it's a unit row and this column is an input field, use submitted data
            val = cell.value
            if matched_sub and r > header_end:
                # Check if this column 'c' is in the config's fields
                is_field = any(f['idx'] == c for f in fields)
                if is_field:
                    val = matched_sub['values'].get(str(c), '')

            display = "" if val is None else str(val)
            if isinstance(val, str) and val.startswith('='): display = ""

            rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
            cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
            
            # Formatting for data rows vs headers
            base_td = 'padding:3px 6px;border:1px solid #d1d5db;overflow:hidden;'
            if r > header_end:
                # Data rows: center by default, add subtle hover
                base_td += 'text-align:center;'
            
            rows_html.append(f'<td{rs_attr}{cs_attr} style="{base_td}{css}">{display}</td>')

        rows_html.append('</tr>')

    html = (
        '<div class="excel-wrapper" style="overflow:auto;max-height:80vh;">'
        '<table class="excel-render-table" '
        'style="border-collapse:collapse;font-size:12px;font-family:\'Calibri\',\'Arial\',sans-serif;">'
        + ''.join(col_parts) +
        '<tbody>' + ''.join(rows_html) + '</tbody>'
        '</table></div>'
    )
    return Markup(html)


def build_v2_stats_table_html(file_blob, metadata, all_values):
    """
    For V2 Stats: render the full Excel template structure and overlay
    data from ALL submissions for that version.

    Parameters
    ----------
    file_blob  : bytes – raw Excel template
    metadata   : dict – parsed metadata for the template version
    all_values : dict – { cell_coord: value } consolidated from all submissions
    """
    if not file_blob:
        return Markup('<p class="text-muted">Không có file Excel gốc.</p>')

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_blob), data_only=True)
    except Exception as e:
        return Markup(f'<p class="text-danger">Lỗi đọc file Excel: {e}</p>')

    sheets_html = []
    from pc06_excel_engine import ExcelEngineV2

    for ws in wb.worksheets:
        # Get region from metadata
        sheet_meta = next((s for s in metadata.get('sheets', []) if s['name'] == ws.title), None)
        if sheet_meta:
            region = sheet_meta.get('activeRenderRegion', {})
            min_row = region.get('r1', 1)
            min_col = region.get('c1', 1)
            max_row = region.get('r2', ws.max_row)
            # Use max of region c2 and ws.max_column to avoid missing rightmost columns
            max_col = max(region.get('c2', ws.max_column), ws.max_column
                          if ws.max_column <= region.get('c2', 0) + 5 else region.get('c2', ws.max_column))
        else:
            max_row, max_col = ExcelEngineV2._get_true_max_row_col(wb, ws)
            min_row, min_col = 1, 1

        # Build merge lookup but only track shadows within our render region
        spans_all, shadows_all = _build_merge_lookup(ws)
        # Filter spans to only those starting within render region
        spans = {k: v for k, v in spans_all.items()
                 if min_row <= k[0] <= max_row and min_col <= k[1] <= max_col}
        # Keep full shadows so we skip covered cells correctly
        shadows = shadows_all

        col_widths = []
        for i in range(min_col, max_col + 1):
            letter = get_column_letter(i)
            w = ws.column_dimensions[letter].width or 8.43
            col_widths.append(max(int(w * 7), 45))

        colgroup = '<colgroup>' + ''.join(f'<col style="width:{w}px">' for w in col_widths) + '</colgroup>'
        rows_html = []

        for r in range(min_row, max_row + 1):
            if ws.row_dimensions[r].hidden:
                continue
            rh = _row_height_px(ws, r)
            rows_html.append(f'<tr style="height:{rh}px">')

            for c in range(min_col, max_col + 1):
                if (r, c) in shadows:
                    continue

                cell = ws.cell(row=r, column=c)
                rowspan, colspan = spans.get((r, c), (1, 1))
                css = _cell_css(cell)

                coord = cell.coordinate
                # Prioritize submitted value, fallback to template value
                val = all_values.get(coord,
                      cell.value if cell.value and not str(cell.value).startswith('=') else '')

                display = '' if val is None else str(val)
                rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''

                base_td = 'padding:3px 6px;border:1px solid #d1d5db;overflow:hidden;'
                rows_html.append(f'<td{rs_attr}{cs_attr} style="{base_td}{css}">{display}</td>')
            rows_html.append('</tr>')

        sheet_title_html = (
            f'<h6 class="fw-bold mt-4 mb-2">'
            f'<i class="fa-solid fa-layer-group me-2"></i>Sheet: {ws.title}</h6>'
        )
        sheet_table = (
            f'<div class="excel-wrapper mb-4" '
            f'style="overflow:auto;max-height:80vh;border:1px solid #eee;border-radius:8px;">'
            f'<table class="excel-render-table" '
            f'style="border-collapse:collapse;font-size:12px;width:max-content;">'
            f'{colgroup}<tbody>{"".join(rows_html)}</tbody></table></div>'
        )
        sheets_html.append(sheet_title_html + sheet_table)

    return Markup(''.join(sheets_html))

