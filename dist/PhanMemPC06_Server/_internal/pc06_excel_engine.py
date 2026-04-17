import openpyxl
from openpyxl.styles import PatternFill
import json
import os
from excel_renderer import is_input_cell

class ExcelEngineV2:
    """
    Advanced Excel-to-Metadata Parser for High-Fidelity Web Rendering.
    """

    @staticmethod
    def parse_template(file_path, input_marker_hex="FFE0F2FE"):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        metadata = {
            "sheets": [],
            "parser_version": "3.0"  # Incremented for region support
        }

        for ws in wb.worksheets:
            # 0. Detect Active Regions
            regions = ExcelEngineV2._detect_active_regions(wb, ws, input_marker_hex)
            
            # Boundaries: [min_col, min_row, max_col, max_row]
            r_box = regions["render"]
            min_col, min_row, max_col, max_row = r_box[0], r_box[1], r_box[2], r_box[3]
            
            sheet_meta = {
                "name": ws.title,
                "activeRenderRegion": {
                    "r1": r_box[1], "c1": r_box[0],
                    "r2": r_box[3], "c2": r_box[2]
                },
                "activeReportRegion": {
                    "r1": regions["report"][1], "c1": regions["report"][0],
                    "r2": regions["report"][3], "c2": regions["report"][2]
                },
                "rowCount": max_row, # For backward compatibility
                "colCount": max_col,
                "rows": [],
                "merges": [],
                "colWidths": {},
                "hiddenRows": [],
                "hiddenCols": []
            }

            # 1. Capture Column Widths (Only for active range)
            for i in range(min_col, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(i)
                width = ws.column_dimensions[col_letter].width
                sheet_meta["colWidths"][col_letter] = width if width is not None else 8.43
                if ws.column_dimensions[col_letter].hidden:
                    sheet_meta["hiddenCols"].append(col_letter)

            # 2. Capture Merges (Filter to active range)
            for merge in ws.merged_cells.ranges:
                # Check if merge intersects with activeRenderRegion
                m_r1, m_c1, m_r2, m_c2 = merge.min_row, merge.min_col, merge.max_row, merge.max_col
                if not (m_r2 < min_row or m_r1 > max_row or m_c2 < min_col or m_c1 > max_col):
                    sheet_meta["merges"].append(str(merge))

            # 3. Iterate Rows & Cells (Only in active range)
            for r in range(min_row, max_row + 1):
                row_dim = ws.row_dimensions[r]
                row_meta = {
                    "index": r,
                    "height": row_dim.height if row_dim.height else 15,
                    "cells": []
                }

                if row_dim.hidden:
                    sheet_meta["hiddenRows"].append(r)

                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell_coord = cell.coordinate

                    # Style Extraction
                    style = ExcelEngineV2._extract_styles(cell)

                    # Identify Input Marker
                    is_input = is_input_cell(cell)

                    # Data Type and Logic
                    cell_type = "input" if is_input else "label"
                    data_type = "text"
                    if cell.data_type == 'n': data_type = "number"
                    elif cell.data_type == 'd': data_type = "date"

                    cell_meta = {
                        "coord": cell_coord,
                        "type": cell_type,
                        "dataType": data_type,
                        "value": cell.value if not is_input else None,
                        "formula": cell.value if cell.data_type == 'f' else None,
                        "style": style,
                        "bindingKey": ExcelEngineV2._get_binding_key(wb, ws, cell)
                    }
                    row_meta["cells"].append(cell_meta)

                sheet_meta["rows"].append(row_meta)

            metadata["sheets"].append(sheet_meta)

        return metadata

    @staticmethod
    def save_logic_to_source(template_name, metadata, base_dir="."):
        """
        Saves the parsed metadata to the project's source directory for transparency and versioning.
        """
        logic_dir = os.path.join(base_dir, "v2_logic_configs")
        if not os.path.exists(logic_dir):
            os.makedirs(logic_dir)

        # Clean name for filename
        from utils import remove_accents
        safe_name = remove_accents(template_name).replace(" ", "_").lower()
        if not safe_name: safe_name = "template_logic"
        file_path = os.path.join(logic_dir, f"{safe_name}.json")

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)

        return file_path

    @staticmethod
    def _extract_styles(cell):
        styles = {}

        # 1. Background Color (Handle PatternFill)
        # Avoid black backgrounds for transparent/default cells
        if cell.fill and isinstance(cell.fill, PatternFill):
            if cell.fill.patternType == 'solid' and hasattr(cell.fill, 'start_color'):
                hex_color = str(cell.fill.start_color.index)
                # Ignore '00000000' (Transparent/Default) and standard White 'FFFFFFFF'
                if hex_color not in ['00000000', 'None', 'FFFFFFFF']:
                    if len(hex_color) == 8: # AARRGGBB
                        styles["backgroundColor"] = f"#{hex_color[2:]}"
                    elif len(hex_color) == 6: # RRGGBB
                        styles["backgroundColor"] = f"#{hex_color}"

        # 2. Font Styles
        if cell.font:
            if cell.font.b: styles["fontWeight"] = "bold"
            if cell.font.i: styles["fontStyle"] = "italic"
            if cell.font.sz: styles["fontSize"] = f"{cell.font.sz}pt"

            # Font Color
            if cell.font.color and hasattr(cell.font.color, 'index'):
                c = str(cell.font.color.index)
                if c not in ['00000000', 'None']:
                    if len(c) == 8: styles["color"] = f"#{c[2:]}"
                    elif len(c) == 6: styles["color"] = f"#{c}"

        # 3. Alignment
        if cell.alignment:
            if cell.alignment.horizontal: styles["textAlign"] = cell.alignment.horizontal
            if cell.alignment.vertical: styles["verticalAlign"] = cell.alignment.vertical

        return styles

    @staticmethod
    def _detect_active_regions(wb, ws, input_marker_hex="FFE0F2FE"):
        """
        Detects meaningful data clusters in the worksheet.
        Returns { "render": [r1, c1, r2, c2], "report": [r1, c1, r2, c2] }
        """
        # 1. Check for Named Ranges (Override)
        # Search for FORM_REGION or REPORT_REGION for this specific sheet
        named_render = None
        named_report = None
        
        for name in wb.defined_names.values():
            # Check if name is scoped to this sheet or global
            # For simplicity, we check if the destination is this worksheet
            try:
                destinations = list(name.destinations)
                for sheet_name, coord in destinations:
                    if sheet_name == ws.title:
                        if name.name == "FORM_REGION": named_render = coord
                        if name.name == "REPORT_REGION": named_report = coord
            except Exception:
                continue

        def coord_to_range(coord):
            # Parse 'A1:B10' or '$A$1:$B$10' to [r1, c1, r2, c2]
            from openpyxl.utils.cell import range_boundaries
            return list(range_boundaries(coord.replace('$', '')))

        if named_render:
            r1, c1, r2, c2 = coord_to_range(named_render)
            return {"render": [r1, c1, r2, c2], "report": coord_to_range(named_report) if named_report else [r1, c1, r2, c2]}

        # 2. Heuristic: Find Meaningful Cells
        meaningful_cells = []
        
        # We scan the sheet up to max_row/max_col (capped at 5000 for safety)
        limit_r = min(ws.max_row, 5000)
        limit_c = min(ws.max_column, 256)

        for r in range(1, limit_r + 1):
            for c in range(1, limit_c + 1):
                cell = ws.cell(row=r, column=c)
                is_meaningful = False
                
                # Check value - IGNORE whitespace only cells
                if cell.value is not None:
                    val_str = str(cell.value).strip()
                    if val_str != '' and val_str != 'None':
                        is_meaningful = True
                
                # Check for input marker (V2 logic)
                if not is_meaningful and is_input_cell(cell):
                    is_meaningful = True
                
                # Check for comment
                if not is_meaningful and cell.comment: is_meaningful = True
                
                if is_meaningful:
                    meaningful_cells.append((r, c))

        if not meaningful_cells:
            return {"render": [1, 1, 1, 1], "report": [1, 1, 1, 1]}

        # 3. Clustering (Very Simple Proximity Grouping)
        # We group cells if they are within 5 rows AND 5 cols of each other
        clusters = []
        for r, c in meaningful_cells:
            added = False
            for cluster in clusters:
                # If cell is close to any cell in the cluster, merge
                # For optimization, we check against cluster boundaries instead of every cell
                if r >= cluster['r1'] - 5 and r <= cluster['r2'] + 5 and \
                   c >= cluster['c1'] - 5 and c <= cluster['c2'] + 5:
                    cluster['r1'] = min(cluster['r1'], r); cluster['r2'] = max(cluster['r2'], r)
                    cluster['c1'] = min(cluster['c1'], c); cluster['c2'] = max(cluster['c2'], c)
                    added = True; break
            if not added:
                clusters.append({'r1': r, 'c1': c, 'r2': r, 'c2': c, 'count': 1})

        # 4. Filter and Select Main Region
        # Select the largest cluster (by bounded area) or simply the bounding box of all if close
        if not clusters: return {"render": [1, 1, 1, 1], "report": [1, 1, 1, 1]}
        
        major = clusters[0]
        for cluster in clusters[1:]:
            if (cluster['r2'] - cluster['r1'] + 1) * (cluster['c2'] - cluster['c1'] + 1) > \
               (major['r2'] - major['r1'] + 1) * (major['c2'] - major['c1'] + 1):
                major = cluster

        # 5. Expand to include Merged Cells
        final_r1, final_c1, final_r2, final_c2 = major['r1'], major['c1'], major['r2'], major['c2']
        for merge_range in ws.merged_cells.ranges:
            m_r1, m_c1, m_r2, m_c2 = merge_range.min_row, merge_range.min_col, merge_range.max_row, merge_range.max_col
            # If merge range intersects our current box, expand
            if not (m_r2 < final_r1 or m_r1 > final_r2 or m_c2 < final_c1 or m_c1 > final_c2):
                final_r1 = min(final_r1, m_r1); final_r2 = max(final_r2, m_r2)
                final_c1 = min(final_c1, m_c1); final_c2 = max(final_c2, m_c2)

        # 6. Final Polish (Add 1-row/col padding if possible)
        final_r1 = max(1, final_r1 - 1)
        final_c1 = max(1, final_c1 - 1)
        final_r2 = min(ws.max_row, final_r2 + 1)
        final_c2 = min(ws.max_column, final_c2 + 1)

        res = [final_c1, final_r1, final_c2, final_r2]
        return {"render": res, "report": res}

    @staticmethod
    def _get_true_max_row_col(wb, ws):
        """
        Utilizes the region detection to return the max dimensions for current callers.
        """
        regions = ExcelEngineV2._detect_active_regions(wb, ws)
        # boundaries order: min_col, min_row, max_col, max_row
        return regions["render"][3], regions["render"][2] 

    @staticmethod
    def _get_binding_key(wb, ws, cell):
        """
        Check if cell belongs to a Named Range.
        """
        cell_coord = f"'{ws.title}'!${openpyxl.utils.get_column_letter(cell.column)}${cell.row}"
        for name in wb.defined_names.values():
            if name.value == cell_coord:
                return name.name
        return cell.coordinate # Fallback to coordinate
