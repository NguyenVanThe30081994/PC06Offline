# Excel Scanner for V2 Reports - Quét cấu trúc Excel
import io
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

def scan_excel_structure(excel_blob):
    """
    Quét cấu trúc file Excel và trả về gợi ý cấu hình.
    Chỉ quét vùng có dữ liệu thực tế (used range), không quét vùng trắng.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_blob), data_only=False)
    ws = wb.active
    
    # Tìm vùng thực tế có dữ liệu (used range)
    # Thay vì dùng max_row/max_column (có thể bao gồm vùng trắng), 
    # ta tìm last row/col có dữ liệu
    used_min_row = ws.min_row
    used_max_row = ws.max_row
    used_min_col = ws.min_column
    used_max_col = ws.max_column
    
    # Nếu worksheet có nhiều dòng trống ở cuối, tìn lại
    # Duyệt từ cuối ngược lên để tìm dòng có dữ liệu
    if used_max_row > used_min_row:
        for row in range(used_max_row, used_min_row, -1):
            has_data = False
            for col in range(used_min_col, used_max_col + 1):
                cell_val = ws.cell(row, col).value
                if cell_val is not None and str(cell_val).strip() != '':
                    has_data = True
                    break
            if has_data:
                used_max_row = row
                break
            else:
                used_max_row = row - 1  # Giảm max_row nếu dòng trống
    
    # Tương tự cho cột
    if used_max_col > used_min_col:
        for col in range(used_max_col, used_min_col, -1):
            has_data = False
            for row in range(used_min_row, used_max_row + 1):
                cell_val = ws.cell(row, col).value
                if cell_val is not None and str(cell_val).strip() != '':
                    has_data = True
                    break
            if has_data:
                used_max_col = col
                break
            else:
                used_max_col = col - 1
    
    result = {
        'total_rows': used_max_row,
        'total_cols': used_max_col,
        'columns': [],  # ['A', 'B', 'C', ...]
        'header_rows': [],
        'data_start_row': 4,  # Default
        'headers': {},  # {row: {col: value}}
        'merged_cells': [],
        'numeric_columns': [],
        'formulas': {},  # {col_letter: formula_type}
        'original_max_row': ws.max_row,  # Lưu lại để debug
        'original_max_col': ws.max_column,
    }
    
    # Get columns (chỉ vùng có dữ liệu)
    for col in range(used_min_col, used_max_col + 1):
        result['columns'].append(get_column_letter(col))
    
    # Scan merged cells
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        result['merged_cells'].append({
            'range': str(merged_range),
            'row': min_row,
            'col_start': get_column_letter(min_col),
            'col_end': get_column_letter(max_col),
            'colspan': max_col - min_col + 1,
            'rowspan': max_row - min_row + 1,
            'value': ws.cell(min_row, min_col).value
        })
    
    # Detect header rows - find first row with text
    header_candidates = []
    for row in range(1, min(10, ws.max_row + 1)):
        has_text = False
        for col in range(1, min(5, ws.max_column + 1)):
            cell_val = ws.cell(row, col).value
            if cell_val and isinstance(cell_val, str) and len(str(cell_val).strip()) > 0:
                has_text = True
                break
        if has_text:
            header_candidates.append(row)
    
    result['header_rows'] = header_candidates
    if header_candidates:
        result['data_start_row'] = max(header_candidates) + 1
    
    # Scan headers
    for row in header_candidates:
        result['headers'][row] = {}
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row, col).value
            if cell_val:
                result['headers'][row][get_column_letter(col)] = str(cell_val).strip()
    
    # Detect numeric columns
    data_row = result['data_start_row']
    if data_row <= ws.max_row:
        numeric_cols = set()
        sample_rows = range(data_row, min(data_row + 5, ws.max_row + 1))
        
        for col in range(1, ws.max_column + 1):
            all_numeric = True
            for row in sample_rows:
                cell_val = ws.cell(row, col).value
                if cell_val is not None and cell_val != '':
                    if not isinstance(cell_val, (int, float)):
                        if isinstance(cell_val, str) and cell_val.startswith('='):
                            continue
                        all_numeric = False
                        break
            if all_numeric:
                numeric_cols.add(get_column_letter(col))
        
        result['numeric_columns'] = list(numeric_cols)
    
    # Detect formulas
    formula_cols = {}
    if data_row <= ws.max_row:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(data_row, col)
            if cell.data_type == 'f':
                formula = str(cell.value)
                if 'SUM' in formula.upper():
                    formula_cols[get_column_letter(col)] = 'SUM'
                elif 'AVG' in formula.upper():
                    formula_cols[get_column_letter(col)] = 'AVG'
                elif '/' in formula and '%' not in formula.upper():
                    formula_cols[get_column_letter(col)] = 'RATIO'
                else:
                    formula_cols[get_column_letter(col)] = 'CUSTOM'
    
    result['formulas'] = formula_cols
    
    return result
