from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify, send_file
from models import db, User
import json
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

excel_builder_bp = Blueprint('excel_builder_bp', __name__)


@excel_builder_bp.route('/admin/excel-builder')
def dashboard():
    """Excel Builder Dashboard"""
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))
    
    return render_template('excel_builder.html')


@excel_builder_bp.route('/admin/excel-builder/create', methods=['POST'])
def create_template():
    """Tạo template mới và chuyển sang chỉnh sửa"""
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))
    
    name = request.form.get('name', 'Biểu mẫu mới')
    rows = int(request.form.get('rows', 10))
    cols = int(request.form.get('cols', 8))
    
    # Create initial grid data
    grid_data = {
        'name': name,
        'rows': rows,
        'cols': cols,
        'headers': {},  # {row_col: value}
        'merges': [],   # ["A1:B1"]
        'column_types': {},  # {col: "number"|"text"|"percent"}
    }
    
    return render_template('excel_builder_edit.html',
                          grid_data=grid_data,
                          template_id=None)


@excel_builder_bp.route('/admin/excel-builder/edit/<template_id>')
def edit_template(template_id):
    """Chỉnh sửa template có sẵn"""
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))
    
    # Load from session (temporary storage)
    template_data = session.get(f'excel_builder_{template_id}')
    if not template_data:
        flash('Không tìm thấy template!', 'danger')
        return redirect(url_for('excel_builder_bp.dashboard'))
    
    return render_template('excel_builder_edit.html',
                          grid_data=template_data,
                          template_id=template_id)


@excel_builder_bp.route('/admin/excel-builder/save', methods=['POST'])
def save_template():
    """Lưu template vào session"""
    if not session.get('is_admin'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    grid_data = json.loads(request.form.get('grid_data', '{}'))
    template_id = request.form.get('template_id')
    
    if not template_id:
        template_id = f"template_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Save to session (temporary)
    session[f'excel_builder_{template_id}'] = grid_data
    
    return jsonify({'status': 'success', 'template_id': template_id})


@excel_builder_bp.route('/admin/excel-builder/export/<template_id>')
def export_template(template_id):
    """Export template ra file Excel"""
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))
    
    # Load template data
    grid_data = session.get(f'excel_builder_{template_id}')
    if not grid_data:
        flash('Không tìm thấy template!', 'danger')
        return redirect(url_for('excel_builder_bp.dashboard'))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    rows = grid_data.get('rows', 10)
    cols = grid_data.get('cols', 8)
    headers = grid_data.get('headers', {})
    column_types = grid_data.get('column_types', {})
    merges = grid_data.get('merges', [])
    
    # Style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for cell_ref, value in headers.items():
        # Parse row_col format: "2_B" means row 2, column B
        row_idx, col_letter = cell_ref.split('_')
        col_idx = ord(col_letter) - ord('A') + 1
        ws.cell(row=int(row_idx), column=col_idx, value=value)
        
        # Style header row
        if int(row_idx) <= 3:  # First 3 rows are headers
            cell = ws.cell(row=int(row_idx), column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    # Apply merges
    for merge_range in merges:
        try:
            ws.merge_cells(merge_range)
        except:
            pass
    
    # Set column types (for data entry rows)
    for col_letter, col_type in column_types.items():
        col_idx = ord(col_letter) - ord('A') + 1
        # Note: Column type is informational in the template
    
    # Set column widths
    for col in range(1, cols + 1):
        ws.column_dimensions[chr(ord('A') + col - 1)].width = 15
    
    # Auto-filter
    if headers:
        max_header_row = max([int(k.split('_')[0]) for k in headers.keys()]) if headers else 1
        ws.auto_filter.ref = f"A1:{chr(ord('A') + cols - 1)}{max_header_row + rows}"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{grid_data.get('name', 'bieumau')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, 
                  mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@excel_builder_bp.route('/admin/excel-builder/preview')
def preview_template():
    """Preview template as HTML table"""
    if not session.get('is_admin'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    grid_data = json.loads(request.args.get('grid_data', '{}'))
    
    rows = grid_data.get('rows', 10)
    cols = grid_data.get('cols', 8)
    headers = grid_data.get('headers', {})
    
    # Build HTML table
    html = '<table class="table table-bordered">'
    
    # Header rows
    for r in range(1, min(4, rows + 1)):
        html += '<tr>'
        for c in range(1, cols + 1):
            col_letter = chr(ord('A') + c - 1)
            cell_key = f"{r}_{col_letter}"
            value = headers.get(cell_key, f"{col_letter}{r}")
            
            is_header = r <= 3
            bg_class = 'bg-primary bg-opacity-10' if is_header else ''
            
            html += f'<td class="{bg_class}" contenteditable="true" data-cell="{cell_key}">{value}</td>'
        html += '</tr>'
    
    # Data rows (empty)
    for r in range(4, rows + 1):
        html += '<tr>'
        for c in range(1, cols + 1):
            col_letter = chr(ord('A') + c - 1)
            html += f'<td contenteditable="true" data-cell="{r}_{col_letter}"></td>'
        html += '</tr>'
    
    html += '</table>'
    
    return jsonify({'html': html})
