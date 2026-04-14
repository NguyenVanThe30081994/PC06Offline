from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify
from models import db, ReportTemplateV2, ReportVersionV2, ReportSubmissionV2, ReportValueV2, ReportAuditV2, User
from pc06_excel_engine import ExcelEngineV2
from utils import normalize_unit_name
import json
import io
import os
from datetime import datetime

reports_v2_bp = Blueprint('reports_v2_bp', __name__)

GLOBAL_UNITS = ['Hệ thống', 'Admin', 'PC06']


def _is_global_user(is_admin, user_unit):
    """Check if user can access all units (admin or special global unit)"""
    return bool(is_admin) or (user_unit in GLOBAL_UNITS)


def _normalize_v2_key(sheet_name, coord):
    return f"{sheet_name}!{coord}"


def _split_v2_key(raw_key):
    key = str(raw_key or '').strip()
    if '!' in key:
        sheet, coord = key.split('!', 1)
        return sheet.strip(), coord.strip().upper()
    return None, key.upper()


def _normalize_text_for_unit_match(value):
    """
    Normalize unit name to a MATCH KEY.
    Extracts the base name: "Công an xã A" -> "a", "UBND xã B" -> "b"
    All these become the same: Công an xã A, xã A, UBND xã A → "xa a"
    """
    if not value:
        return ""
    
    # Normalize first
    txt = normalize_unit_name(value or '')
    txt = ' '.join(str(txt).split())
    
    # If empty after normalize, use the original
    if not txt:
        txt = normalize_unit_name(str(value))
    
    return txt


def _extract_unit_key(name):
    """
    Extract the CORE unit key from unit name.
    "Công an xã An Tường" → "an tuong"
    "UBND xã B" → "b"
    "xã A" → "a"
    """
    if not name:
        return ""
    import unicodedata
    # Lowercase + remove accents
    n = str(name).lower().strip()
    n = unicodedata.normalize('NFKD', n).encode('ascii', 'ignore').decode('utf-8')
    
    # Remove common prefixes but KEEP the unit name
    prefixes = [
        "cong an ", "cong an phuong ", "cong an xa ", "cong an huyen ", "cong an tinh ",
        "ubnd ", "ubnd xa ", "ubnd phuong ", "ubnd huyen ",
        "xa ", "phuong ", "huyen ", "thi tran ", "thanh pho "
    ]
    for p in prefixes:
        if n.startswith(p):
            n = n[len(p):].strip()
            break
    
    # Clean up
    n = ' '.join(n.split())
    return n


def _row_matches_unit(ws, row_idx, min_col, max_col, norm_user_unit):
    """
    Match row with user unit - SMART matching.
    Matches if normalized names match OR if core keys match.
    "Công an xã A" matches "Xã A", "UBND xã A", etc.
    """
    # Extract core key from user unit
    user_key = _extract_unit_key(norm_user_unit)
    
    for c in range(min_col, max_col + 1):
        cell_val = ws.cell(row=row_idx, column=c).value
        if cell_val is None:
            continue
        
        cell_str = str(cell_val).strip()
        norm_cell = _normalize_text_for_unit_match(cell_str)
        cell_key = _extract_unit_key(cell_str)
        
        if not norm_cell:
            continue
        
        # Match 1: Exact normalized match
        if norm_cell == norm_user_unit:
            return True, c
        
        # Match 2: Core key match (e.g., "an tuong" matches "an tuong")
        if user_key and cell_key and (user_key == cell_key):
            return True, c
        
        # Match 3: User key is contained in cell (for partial matches like "Công an xã An Tường" in header)
        if user_key and cell_key and (user_key in cell_key or cell_key in user_key):
            return True, c
    
    return False, None


def _find_unit_rows_and_col(ws, min_row, max_row, min_col, max_col, user_unit):
    """
    Find rows matching user unit.
    - If user_unit is empty → return empty (no access without proper unit)
    - Strict exact match only
    """
    norm_user_unit = _normalize_text_for_unit_match(user_unit)
    matched_rows = []
    matched_col = None
    
    # If no unit assigned → no access to any row
    if not norm_user_unit:
        return matched_rows, matched_col

    for r in range(min_row, max_row + 1):
        is_match, col_idx = _row_matches_unit(ws, r, min_col, max_col, norm_user_unit)
        if is_match:
            matched_rows.append(r)
            if matched_col is None and col_idx is not None:
                matched_col = col_idx
    return matched_rows, matched_col


def _is_editable_by_row_context(cell, col_idx, unit_col):
    if unit_col is not None and col_idx <= unit_col:
        return False
    if isinstance(cell.value, str):
        raw = cell.value.strip()
        if raw == '':
            return True
        if raw.startswith('='):
            return False
        return False
    if cell.value is None:
        return True
    if isinstance(cell.value, (int, float)):
        return True
    return False


def _get_sheet_region(meta_data, ws, wb):
    sheet_meta = next((s for s in meta_data.get('sheets', []) if s.get('name') == ws.title), None)
    if sheet_meta:
        region = sheet_meta.get('activeRenderRegion', {})
        min_row = region.get('r1', 1)
        min_col = region.get('c1', 1)
        max_row = region.get('r2', ws.max_row)
        max_col = region.get('c2', ws.max_column)
    else:
        max_row, max_col = ExcelEngineV2._get_true_max_row_col(wb, ws)
        min_row, min_col = 1, 1
    return min_row, min_col, max_row, max_col


def _collect_allowed_input_keys(wb, meta_data, user_unit, is_admin):
    from excel_renderer import is_input_cell

    is_global = _is_global_user(is_admin, user_unit)
    allowed_keys = set()

    for ws in wb.worksheets:
        min_row, min_col, max_row, max_col = _get_sheet_region(meta_data, ws, wb)
        unit_rows, unit_col = _find_unit_rows_and_col(ws, min_row, max_row, min_col, max_col, user_unit)

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                coord = cell.coordinate
                key = _normalize_v2_key(ws.title, coord)

                if is_global:
                    if is_input_cell(cell):
                        allowed_keys.add(key)
                    continue

                if r in unit_rows and _is_editable_by_row_context(cell, c, unit_col):
                    allowed_keys.add(key)

    return allowed_keys


@reports_v2_bp.route('/reports-v2')
def dashboard():
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))
    
    templates = ReportTemplateV2.query.order_by(ReportTemplateV2.created_at.desc()).all()
    is_admin = session.get('is_admin', False)
    
    # Only use mobile template if explicitly requested via ?mobile=1
    # This ensures Desktop browsers always get Desktop template
    if request.args.get('mobile') == '1':
        return render_template('reports_v2_dashboard_mobile.html', templates=templates, is_admin=is_admin)
    
    return render_template('reports_v2_dashboard.html', templates=templates, is_admin=is_admin)


@reports_v2_bp.route('/reports-v2/upload', methods=['POST'])
def upload_template():
    if not session.get('is_admin'):
        return jsonify({"error": "Unauthorized"}), 403

    file = request.files.get('template_excel')
    name = request.form.get('name', 'Báo cáo Mới')
    is_daily = request.form.get('is_daily') == 'true' or 'is_daily' in request.form

    if not file or not file.filename:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        file_content = file.read()
        
        # Sanitize filename to avoid encoding errors
        import unicodedata
        safe_filename = unicodedata.normalize('NFD', file.filename)
        safe_filename = ''.join(c for c in safe_filename if unicodedata.category(c) != 'Mn')
        safe_filename = ''.join(c if c.isalnum() or c in '._-' else '_' for c in safe_filename)
        
        temp_path = os.path.join("tmp", safe_filename)
        os.makedirs("tmp", exist_ok=True)
        with open(temp_path, "wb") as f:
            f.write(file_content)

        metadata = ExcelEngineV2.parse_template(temp_path)

        template = ReportTemplateV2.query.filter_by(name=name).first()
        is_new = False
        if not template:
            template = ReportTemplateV2(
                name=name,
                created_by=session.get('fullname'),
                is_daily=is_daily
            )
            db.session.add(template)
            db.session.flush()
            is_new = True
        else:
            template.is_daily = is_daily

        new_version = ReportVersionV2(
            template_id=template.id,
            version_tag=datetime.now().strftime("%Y%m%d%H%M"),
            metadata_json=json.dumps(metadata, ensure_ascii=False),
            excel_file_blob=file_content,
            is_published=True
        )
        db.session.add(new_version)

        ExcelEngineV2.save_logic_to_source(name, metadata)

        db.session.commit()

        if is_new:
            from utils import push_global_notif
            push_global_notif("Biểu mẫu mới", f"Vừa có biểu mẫu mới: {name}", f"/reports-v2/render/{template.id}", exclude_uid=session['uid'])

        if os.path.exists(temp_path):
            os.remove(temp_path)

        return jsonify({"success": True, "template_id": template.id, "version_id": new_version.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@reports_v2_bp.route('/reports-v2/edit/<int:tid>', methods=['GET', 'POST'])
def edit_template(tid):
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))

    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        flash('Không tìm thấy biểu mẫu V2!', 'danger')
        return redirect(url_for('reports_v2_bp.dashboard'))

    if request.method == 'POST':
        template.name = request.form.get('name', template.name)
        template.description = request.form.get('description', template.description)
        template.is_daily = 'is_daily' in request.form

        file = request.files.get('template_excel')
        if file and file.filename:
            try:
                file_content = file.read()
                temp_path = os.path.join("tmp", file.filename)
                os.makedirs("tmp", exist_ok=True)
                with open(temp_path, "wb") as f:
                    f.write(file_content)

                metadata = ExcelEngineV2.parse_template(temp_path)
                ReportVersionV2.query.filter_by(template_id=tid, is_published=True).update({'is_published': False})

                new_version = ReportVersionV2(
                    template_id=tid,
                    version_tag=datetime.now().strftime("%Y%m%d%H%M"),
                    metadata_json=json.dumps(metadata, ensure_ascii=False),
                    excel_file_blob=file_content,
                    is_published=True
                )
                db.session.add(new_version)
                ExcelEngineV2.save_logic_to_source(template.name, metadata)

                if os.path.exists(temp_path):
                    os.remove(temp_path)

                flash('Đã cập nhật file Excel và tạo phiên bản mới!', 'info')
            except Exception as e:
                db.session.rollback()
                flash(f'Lỗi xử lý file: {str(e)}', 'danger')
                return redirect(url_for('reports_v2_bp.edit_template', tid=tid))

        db.session.commit()
        flash('Đã cập nhật biểu mẫu V2!', 'success')
        return redirect(url_for('reports_v2_bp.dashboard'))

    versions = ReportVersionV2.query.filter_by(template_id=tid).order_by(ReportVersionV2.created_at.desc()).all()
    
    # Check if mobile device - redirect to dashboard if so
    user_agent = request.headers.get('User-Agent', '').lower()
    is_mobile_request = 'mobile' in user_agent or 'android' in user_agent or 'iphone' in user_agent
    
    if is_mobile_request:
        flash('Vui lòng sử dụng máy tính để chỉnh sửa biểu mẫu V2.', 'warning')
        return redirect(url_for('reports_v2_bp.dashboard'))
    
    return render_template('reports_v2_edit.html', template=template, versions=versions)


@reports_v2_bp.route('/reports-v2/config/<int:tid>')
def config_template(tid):
    """Giao diện cấu hình cột cho template"""
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))
    
    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        flash('Không tìm thấy biểu mẫu!', 'danger')
        return redirect(url_for('reports_v2_bp.dashboard'))
    
    version = ReportVersionV2.query.filter_by(template_id=tid, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
    if not version or not version.excel_file_blob:
        flash('Chưa có file Excel! Vui lòng upload trước.', 'warning')
        return redirect(url_for('reports_v2_bp.edit_template', tid=tid))
    
    # Quét cấu trúc Excel
    from pc06_excel_scanner import scan_excel_structure
    detected = scan_excel_structure(version.excel_file_blob)
    
    # Load existing config
    try:
        existing_config = json.loads(version.metadata_json or '{}')
    except:
        existing_config = {}
    
    column_configs = existing_config.get('column_configs', {})
    header_groups = existing_config.get('header_groups', [])
    data_start_row = existing_config.get('data_start_row', detected.get('data_start_row', 4))
    unit_column = existing_config.get('unit_column', 'B')
    
    return render_template('reports_v2_config.html',
                          template=template,
                          detected=detected,
                          column_configs=column_configs,
                          header_groups=header_groups,
                          data_start_row=data_start_row,
                          unit_column=unit_column)


@reports_v2_bp.route('/reports-v2/config/<int:tid>', methods=['POST'])
def save_config(tid):
    """Lưu cấu hình cột cho template"""
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))
    
    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        flash('Không tìm thấy biểu mẫu!', 'danger')
        return redirect(url_for('reports_v2_bp.dashboard'))
    
    version = ReportVersionV2.query.filter_by(template_id=tid, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
    if not version:
        flash('Không có phiên bản!', 'danger')
        return redirect(url_for('reports_v2_bp.dashboard'))
    
    try:
        # Load existing metadata
        try:
            metadata = json.loads(version.metadata_json or '{}')
        except:
            metadata = {}
        
        # Build column configs from form data
        column_configs = {}
        all_columns = request.form.getlist('columns') if request.form.get('columns') else []
        
        # If no columns in form, try to get from detected
        if not all_columns:
            # Get all column configs from form
            for key in request.form.keys():
                if key.startswith('is_numeric_') or key.startswith('is_percent_') or key.startswith('is_formula_') or key.startswith('is_sortable_'):
                    col = key.split('_', 2)[-1]
                    if col not in column_configs:
                        column_configs[col] = {}
        
        for col in column_configs:
            column_configs[col] = {
                'is_numeric': f'is_numeric_{col}' in request.form,
                'is_percent': f'is_percent_{col}' in request.form,
                'is_formula': f'is_formula_{col}' in request.form,
                'is_sortable': f'is_sortable_{col}' in request.form,
            }
        
        metadata['column_configs'] = column_configs
        metadata['data_start_row'] = int(request.form.get('data_start_row', 4))
        metadata['unit_column'] = request.form.get('unit_column', 'B')
        metadata['config_version'] = '2.0'
        
        version.metadata_json = json.dumps(metadata, ensure_ascii=False)
        db.session.commit()
        
        flash('Đã lưu cấu hình V2!', 'success')
        return redirect(url_for('reports_v2_bp.dashboard'))
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi: {str(e)}', 'danger')
        return redirect(url_for('reports_v2_bp.config_template', tid=tid))


@reports_v2_bp.route('/reports-v2/delete/<int:tid>', methods=['POST'])
def delete_template(tid):
    if not session.get('is_admin'):
        return jsonify({"error": "Unauthorized"}), 403

    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        return jsonify({"error": "Not found"}), 404

    try:
        versions = ReportVersionV2.query.filter_by(template_id=tid).all()
        for v in versions:
            for sub in ReportSubmissionV2.query.filter_by(version_id=v.id).all():
                ReportValueV2.query.filter_by(submission_id=sub.id).delete()
                ReportAuditV2.query.filter_by(submission_id=sub.id).delete()
                db.session.delete(sub)
            db.session.delete(v)
        db.session.delete(template)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@reports_v2_bp.route('/reports-v2/render/<int:tid>')
def render_report(tid):
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        return "Template Not Found", 404

    version = ReportVersionV2.query.filter_by(template_id=tid, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
    if not version:
        return "No published version found", 404

    if not version.excel_file_blob:
        return "No Excel file stored for this version", 400

    submission = ReportSubmissionV2.query.filter_by(version_id=version.id, user_id=session['uid'], status='draft').first()
    existing_values = {}
    if submission:
        for val in submission.values:
            existing_values[val.cell_key] = val.value

    user_unit = session.get('unit', session.get('unit_area', ''))
    is_admin = session.get('is_admin', False)
    is_global = _is_global_user(is_admin, user_unit)

    from excel_renderer import _build_merge_lookup, _col_widths_px, _row_height_px, _cell_css, is_input_cell
    import openpyxl as _opx

    try:
        wb = _opx.load_workbook(io.BytesIO(version.excel_file_blob), data_only=True)
    except Exception as e:
        return f"Error loading Excel: {e}", 500

    meta_data = {}
    try:
        meta_data = json.loads(version.metadata_json or '{}')
    except Exception:
        meta_data = {}

    sheets_html = []
    for ws in wb.worksheets:
        spans, shadows = _build_merge_lookup(ws)
        col_widths = _col_widths_px(ws)
        colgroup = '<colgroup>' + ''.join(f'<col style="width:{w}px">' for w in col_widths) + '</colgroup>'

        min_row, min_col, max_row, max_col = _get_sheet_region(meta_data, ws, wb)
        unit_rows, unit_col = _find_unit_rows_and_col(ws, min_row, max_row, min_col, max_col, user_unit)
        
        # Find all unit header rows in sheet
        all_unit_keys = set()
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell_val = ws.cell(row=r, column=c).value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    norm_cell = _normalize_text_for_unit_match(cell_str)
                    cell_key = _extract_unit_key(norm_cell)
                    if cell_key:
                        all_unit_keys.add((r, cell_key))
                        break
        
        # Determine show/hide for each row
        first_unit_row = min([r for r, k in all_unit_keys]) if all_unit_keys else None
        user_rows_set = set(unit_rows)
        other_unit_rows = {r for r, k in all_unit_keys if r not in user_rows_set}
        
        # Find end of user's data section
        user_data_end = max_row
        if unit_rows:
            user_first = min(unit_rows)
            for r, key in sorted(all_unit_keys):
                if r > user_first:
                    user_data_end = r
                    break

        rows_html = []
        for r in range(min_row, max_row + 1):
            if ws.row_dimensions[r].hidden:
                continue

            # Show user's unit + headers, hide other units
            if not is_global and unit_rows and all_unit_keys:
                user_first = min(unit_rows)
                # Show user's unit rows
                if r in unit_rows:
                    pass
                # Hide other unit headers
                elif r in other_unit_rows:
                    continue
                # Show data between user's unit and next unit  
                elif r > user_first and r <= user_data_end:
                    pass
                # Show headers before first unit in sheet
                elif r < first_unit_row:
                    pass
                else:
                    continue

            rh = _row_height_px(ws, r)
            row_parts = [f'<tr style="height:{rh}px">']
            for c in range(min_col, max_col + 1):
                if (r, c) in shadows:
                    continue

                cell = ws.cell(row=r, column=c)
                rowspan, colspan = spans.get((r, c), (1, 1))
                css = _cell_css(cell)

                if is_global:
                    is_input = is_input_cell(cell)
                else:
                    is_input = (r in unit_rows) and _is_editable_by_row_context(cell, c, unit_col)

                coord = cell.coordinate
                key = _normalize_v2_key(ws.title, coord)
                rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                base = 'padding:3px 6px;border:1px solid #d1d5db;overflow:hidden;box-sizing:border-box;'
                if is_input:
                    base += 'background-color:#e0f2fe;'
                full_css = base + css
                td = f'<td{rs_attr}{cs_attr} style="{full_css}">'

                if is_input:
                    val = existing_values.get(key, existing_values.get(coord, ''))
                    safe_val = str(val).replace('"', '&quot;')
                    inner = (
                        f'<input type="text" class="grid-input" '
                        f'data-key="{key}" data-coord="{coord}" '
                        f'value="{safe_val}" onchange="markDirty()" '
                        f'style="width:100%;height:100%;border:none;background:transparent;padding:2px;font-size:inherit;">'
                    )
                else:
                    raw = cell.value
                    if isinstance(raw, str) and raw.startswith('='):
                        raw = ''
                    inner = '' if raw is None else str(raw)

                row_parts.append(f'{td}{inner}</td>')
            row_parts.append('</tr>')
            rows_html.append(''.join(row_parts))

        table_html = (
            f'<table class="excel-render-table" '
            f'style="border-collapse:collapse;font-size:12px;width:100%;table-layout:fixed;'
            f'font-family:Calibri,Arial,sans-serif;min-width:1000px;">'
            f'{colgroup}<tbody>{"".join(rows_html)}</tbody></table>'
        )

        from markupsafe import Markup
        sheets_html.append({'name': ws.title, 'html': Markup(table_html)})

    from models import ReportConfig
    v2_templates = ReportTemplateV2.query.all()
    v1_configs = ReportConfig.query.all()
    return render_template(
        'reports_v2_render.html',
        template=template,
        version=version,
        sheets_html=sheets_html,
        is_admin=is_admin,
        user_unit=user_unit,
        v2_templates=v2_templates,
        v1_configs=v1_configs,
        form_type='v2'
    )


@reports_v2_bp.route('/reports-v2/submit', methods=['POST'])
def submit_data():
    if not session.get('uid'):
        return jsonify({"error": "Unauthorized"}), 403

    data = request.json or {}
    version_id = data.get('version_id')
    values = data.get('values', {})

    if not version_id:
        return jsonify({'success': False, 'message': 'Thiếu thông tin phiên bản mẫu.'}), 400

    version = db.session.get(ReportVersionV2, version_id)
    if not version or not version.excel_file_blob:
        return jsonify({'success': False, 'message': 'Phiên bản mẫu không hợp lệ.'}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(version.excel_file_blob), data_only=True)

        meta_data = {}
        try:
            meta_data = json.loads(version.metadata_json or '{}')
        except Exception:
            meta_data = {}

        user_unit = session.get('unit_area', session.get('unit', 'PC06'))
        is_admin = session.get('is_admin', False)

        allowed_prefixed_keys = _collect_allowed_input_keys(wb, meta_data, user_unit, is_admin)

        normalized_payload = {}
        for raw_key, val in values.items():
            sheet_name, coord = _split_v2_key(raw_key)
            resolved_key = None

            if sheet_name:
                candidate = _normalize_v2_key(sheet_name, coord)
                if candidate in allowed_prefixed_keys:
                    resolved_key = candidate
            else:
                matched = [k for k in allowed_prefixed_keys if k.endswith(f'!{coord}')]
                if len(matched) == 1:
                    resolved_key = matched[0]

            if not resolved_key:
                return jsonify({'success': False, 'message': f'Lỗi bảo mật: Ô {raw_key} không thuộc quyền quản lý của đơn vị bạn!'}), 403

            normalized_payload[resolved_key] = val

        submission = ReportSubmissionV2.query.filter_by(version_id=version_id, user_id=session['uid'], status='draft').first()

        if not submission:
            submission = ReportSubmissionV2(
                version_id=version_id,
                user_id=session['uid'],
                org_unit=session.get('unit_area', session.get('unit', 'PC06'))
            )
            db.session.add(submission)
            db.session.flush()

        existing_rows = ReportValueV2.query.filter_by(submission_id=submission.id).all()
        existing_map = {row.cell_key: row for row in existing_rows}

        for key, val in normalized_payload.items():
            existing_val = existing_map.get(key)
            old_val = existing_val.value if existing_val else None

            if str(old_val) != str(val):
                db.session.add(ReportAuditV2(
                    submission_id=submission.id,
                    user_id=session['uid'],
                    cell_key=key,
                    old_value=old_val,
                    new_value=val
                ))

                if existing_val:
                    existing_val.value = val
                else:
                    db.session.add(ReportValueV2(submission_id=submission.id, cell_key=key, value=val))

        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@reports_v2_bp.route('/reports-v2/export/<int:sid>')
def export_submission(sid):
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    submission = db.session.get(ReportSubmissionV2, sid)
    if not submission:
        return "Submission Not Found", 404

    version = submission.version
    if not version.excel_file_blob:
        return "No original template file found", 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(version.excel_file_blob))

        ws_by_title = {ws.title: ws for ws in wb.worksheets}
        for val in submission.values:
            sheet_name, coord = _split_v2_key(val.cell_key)
            if sheet_name and sheet_name in ws_by_title:
                try:
                    ws_by_title[sheet_name][coord].value = val.value
                except Exception:
                    pass
            else:
                for ws in wb.worksheets:
                    try:
                        ws[coord].value = val.value
                    except Exception:
                        pass

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Report_{submission.org_unit}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        from flask import send_file
        return send_file(output, as_attachment=True, download_name=filename)
    except Exception as e:
        return str(e), 500


@reports_v2_bp.route('/reports-v2/submission/<int:sub_id>')
def review_submission(sub_id):
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    submission = db.session.get(ReportSubmissionV2, sub_id)
    if not submission:
        flash('Không tìm thấy bản nộp!', 'danger')
        return redirect(url_for('forms_bp.stats'))

    version = submission.version
    template = version.template

    existing_values = {val.cell_key: val.value for val in submission.values}

    from excel_renderer import _build_merge_lookup, _row_height_px, _cell_css, is_input_cell
    import openpyxl as _opx

    try:
        wb = _opx.load_workbook(io.BytesIO(version.excel_file_blob), data_only=True)
    except Exception as e:
        return f"Error loading Excel: {e}", 500

    sheets_html = []
    meta_data = {}
    try:
        meta_data = json.loads(version.metadata_json or '{}')
    except Exception:
        meta_data = {}

    for ws in wb.worksheets:
        min_row, min_col, max_row, max_col = _get_sheet_region(meta_data, ws, wb)
        spans, shadows = _build_merge_lookup(ws)

        col_widths = []
        for i in range(min_col, max_col + 1):
            letter = _opx.utils.get_column_letter(i)
            w = ws.column_dimensions[letter].width or 8.43
            col_widths.append(max(int(w * 7), 45))

        colgroup = '<colgroup>' + ''.join(f'<col style="width:{w}px">' for w in col_widths) + '</colgroup>'
        rows_html = []

        for r in range(min_row, max_row + 1):
            if ws.row_dimensions[r].hidden:
                continue
            rh = _row_height_px(ws, r)
            row_parts = [f'<tr style="height:{rh}px">']

            for c in range(min_col, max_col + 1):
                if (r, c) in shadows:
                    continue

                cell = ws.cell(row=r, column=c)
                rowspan, colspan = spans.get((r, c), (1, 1))
                css = _cell_css(cell)

                coord = cell.coordinate
                key = _normalize_v2_key(ws.title, coord)
                val = existing_values.get(key, existing_values.get(coord, cell.value if cell.value and not str(cell.value).startswith('=') else ''))

                rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                bg = 'background-color:#f0f9ff;' if is_input_cell(cell) else ''
                td = f'<td{rs_attr}{cs_attr} style="padding:3px 6px;border:1px solid #d1d5db;{bg}{css}">'
                row_parts.append(f'{td}{val}</td>')

            row_parts.append('</tr>')
            rows_html.append(''.join(row_parts))

        sheets_html.append({
            'name': ws.title,
            'html': f'<table class="excel-render-table" style="border-collapse:collapse;font-size:12px;width:100%;table-layout:fixed;font-family:Calibri,Arial,sans-serif;min-width:1000px;">{colgroup}<tbody>{"".join(rows_html)}</tbody></table>'
        })

    return render_template('reports_v2_review.html', submission=submission, template=template, sheets=sheets_html)
