from flask import Blueprint, render_template, request, session, redirect, url_for, flash
from models import db, ReportConfig, ReportData, User, ReportTemplateV2, ReportVersionV2, ReportSubmissionV2, ReportValueV2, AppRole
import json, io
from datetime import datetime
from utils import remove_accents, log_action

forms_bp = Blueprint('forms_bp', __name__)

@forms_bp.route('/admin-forms', methods=['GET', 'POST'])
def admin_forms():
    if not session.get('is_admin'): return redirect(url_for('auth_bp.login'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description', '')
        is_daily = 'is_daily' in request.form
        header_start = int(request.form.get('header_start', 1))
        header_rows = int(request.form.get('header_rows', 1))
        file = request.files.get('template_excel')
        fid = remove_accents(name).replace(' ', '_')
        
        file_blob = None
        config_json = request.form.get('config_json', '[]')
        
        if file and file.filename:
            file_blob = file.read()
            try:
                import openpyxl as _openpyxl
                wb = _openpyxl.load_workbook(io.BytesIO(file_blob))
                ws = wb.active

                # Build merge map: (row, col) -> value of the top-left cell of that merge
                # Only cells inside an actual merged range are in this map.
                merge_map = {}
                for merge_range in ws.merged_cells.ranges:
                    top_val = ws.cell(merge_range.min_row, merge_range.min_col).value
                    for r in range(merge_range.min_row, merge_range.max_row + 1):
                        for c in range(merge_range.min_col, merge_range.max_col + 1):
                            merge_map[(r, c)] = top_val

                num_cols = ws.max_column

                # --- GROUP DETECTION via actual merge ranges only (no ffill) ---
                # Parent rows = header_start .. header_start + header_rows - 2  (1-indexed)
                group_label_map = {}
                for col_0 in range(num_cols):
                    col_1 = col_0 + 1
                    group_parts = []
                    for parent_row in range(header_start, header_start + header_rows - 1):
                        # Only use the value if the cell is inside a real merge range
                        if (parent_row, col_1) in merge_map:
                            raw = merge_map[(parent_row, col_1)]
                        else:
                            # Cell has its own direct value (single-column parent header)
                            raw = ws.cell(parent_row, col_1).value
                            # If it's populated AND not already the label of this exact column
                            # only, treat it as a group if more than one col shares this value
                            # via merging. Single cells with their own value = no group.
                            raw = None  # standalone parent → no group
                        if raw and str(raw).strip() and str(raw).lower() != 'nan':
                            v = str(raw).strip()
                            if not group_parts or group_parts[-1] != v:
                                group_parts.append(v)
                    group_label_map[col_0] = ' > '.join(group_parts)

                # Last header row = the actual field labels
                last_row_idx = header_start + header_rows - 1  # 1-indexed
                fields = []
                for col_0 in range(num_cols):
                    col_1 = col_0 + 1
                    if (last_row_idx, col_1) in merge_map:
                        cell_val = merge_map[(last_row_idx, col_1)]
                    else:
                        cell_val = ws.cell(last_row_idx, col_1).value
                    cell_val = str(cell_val).strip() if cell_val else ''
                    if not cell_val or cell_val.lower() == 'nan':
                        if group_label_map.get(col_0):
                            cell_val = group_label_map[col_0]
                        else:
                            continue

                    group_label = group_label_map.get(col_0, '')
                    combined = (cell_val + ' ' + group_label).lower()
                    is_num = any(k in combined for k in ["số", "tỷ lệ", "tổng", "năm", "tháng", "quý", "%", "tiền"])
                    fields.append({
                        "idx": col_0 + 1,
                        "label": cell_val,
                        "group": group_label,
                        "type": "number" if is_num else "text",
                        "is_perc": "%" in cell_val or "tỷ lệ" in cell_val.lower()
                                   or "%" in group_label or "tỷ lệ" in group_label.lower(),
                        "is_visible": True
                    })
                if not fields:
                    flash('Không tìm thấy cột dữ liệu nào trong vùng tiêu đề đã chọn!', 'warning')
                else:
                    config_json = json.dumps(fields, ensure_ascii=False)
                    log_action(session['uid'], session['fullname'], "Cấu hình biểu mẫu thành công", "Biểu mẫu", name)
            except Exception as e:
                flash(f'Lỗi phân tích file Excel: {str(e)}', 'danger')
                return redirect(url_for('forms_bp.admin_forms'))
        
        # Improvement: Ensure fid is safe and unique-ish if name is empty
        if not name: 
            flash('Tên biểu mẫu không được để trống!', 'danger')
            return redirect(url_for('forms_bp.admin_forms'))
            
        fid = remove_accents(name).strip().replace(' ', '_')
        fid = "".join([c for c in fid if c.isalnum() or c == '_'])
        
        if not fid: fid = "form_" + str(int(datetime.now().timestamp()))

        existing = db.session.get(ReportConfig, fid)
        if existing:
            existing.name = name
            existing.is_daily = is_daily
            existing.header_start = header_start
            existing.header_rows = header_rows
            existing.config_json = config_json
            if file_blob: existing.file_blob = file_blob
        else:
            db.session.add(ReportConfig(
                id=fid, name=name, description=description, is_daily=is_daily,
                header_start=header_start, header_rows=header_rows,
                config_json=config_json, file_blob=file_blob,
                author_name=session['fullname']
            ))
        db.session.commit()
        flash('Đã lưu cấu hình biểu mẫu! Bạn có thể tinh chỉnh các cột bên dưới.', 'success')
        return redirect(url_for('forms_bp.config_form', fid=fid))

    # Fetch both versions
    configs = ReportConfig.query.all()
    v2_templates = ReportTemplateV2.query.all()
    
    # Pre-calculate counts for V1
    counts = {}
    for c in configs:
        counts[c.id] = ReportData.query.filter_by(report_id=c.id).count()

    # Pre-calculate counts for V2 (Total submissions across all versions)
    v2_counts = {}
    for t in v2_templates:
        sub_count = 0
        for v in t.versions:
            sub_count += len(v.submissions)
        v2_counts[t.id] = sub_count

    form_type = request.args.get('form_type', '')
    
    return render_template('forms_dashboard.html', 
                           configs=configs, 
                           v2_templates=v2_templates,
                           counts=counts,
                           v2_counts=v2_counts,
                           form_type=form_type)

@forms_bp.route('/admin-forms/config/<fid>', methods=['GET', 'POST'])
def config_form(fid):
    if not session.get('is_admin'): return redirect(url_for('auth_bp.login'))
    
    config = db.session.get(ReportConfig, fid)
    if not config:
        flash('Không tìm thấy biểu mẫu!', 'danger')
        return redirect(url_for('forms_bp.admin_forms'))
        
    fields = []
    if config.config_json:
        fields = json.loads(config.config_json)
        
    if request.method == 'POST':
        # Update fields from form
        new_fields = []
        for f in fields:
            idx = str(f['idx'])
            f['label'] = request.form.get(f'label_{idx}', f['label'])
            f['group'] = request.form.get(f'group_{idx}', f.get('group', ''))
            f['type'] = request.form.get(f'type_{idx}', 'text')
            f['is_perc'] = f'perc_{idx}' in request.form
            f['is_visible'] = f'visible_{idx}' in request.form
            f['order'] = int(request.form.get(f'order_{idx}', f['idx']))
            new_fields.append(f)
            
        # Re-sort by order
        new_fields = sorted(new_fields, key=lambda x: x.get('order', x['idx']))
        
        config.config_json = json.dumps(new_fields, ensure_ascii=False)
        db.session.commit()
        flash('Đã cập nhật cấu hình biểu mẫu V1!', 'success')
        return redirect(url_for('forms_bp.admin_forms'))
        
    return render_template('forms_config.html', config=config, fields=fields)

@forms_bp.route('/admin-forms/delete/<fid>')
def delete_form(fid):
    if not session.get('is_admin'): return redirect(url_for('auth_bp.login'))
    c = db.session.get(ReportConfig, fid)
    if c:
        ReportData.query.filter_by(report_id=fid).delete()
        db.session.delete(c)
        db.session.commit()
        flash('Đã xóa biểu mẫu!', 'success')
    return redirect(url_for('forms_bp.admin_forms'))

@forms_bp.route('/input', methods=['GET', 'POST'])
def input_data():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    rid = request.args.get('rid')
    configs = ReportConfig.query.all()
    if not rid:
        return render_template('input.html', configs=configs, active=None, fields=[])
        
    active = db.session.get(ReportConfig, rid)
    if not active:
        flash('Biểu mẫu không tồn tại!', 'warning')
        return redirect(url_for('forms_bp.input_data'))
        
    fields = []
    try:
        if active.config_json:
            fields = json.loads(active.config_json)
    except: pass

    if request.method == 'POST':
        data = request.form.to_dict()
        try:
            # Check if this user already submitted today (if it's a daily report)
            if active.is_daily:
                today = datetime.now().date()
                exists = ReportData.query.filter_by(report_id=rid, user_id=session['uid'], report_date=today).first()
                if exists:
                    flash('Bạn đã gửi báo cáo này trong hôm nay rồi!', 'warning')
                    return redirect(url_for('forms_bp.input_data', rid=rid))

            new_entry = ReportData(
                report_id=rid, 
                user_id=session['uid'], 
                data_json=json.dumps(data, ensure_ascii=False), 
                report_date=datetime.now().date()
            )
            db.session.add(new_entry)
            db.session.commit()
            flash('Gửi dữ liệu báo cáo thành công!', 'success')
            log_action(session['uid'], session['fullname'], "Gửi báo cáo", "Biểu mẫu", active.name)
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi khi lưu dữ liệu: {e}', 'danger')
            
        return redirect(url_for('forms_bp.input_data', rid=rid))
        
    return render_template('input.html', configs=configs, active=active, fields=fields, form_type='v1')

@forms_bp.route('/stats')
def stats():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    rid = request.args.get('rid')
    is_v2 = request.args.get('v2') == '1'
    form_type = request.args.get('form_type')
    
    # Logic to determine initial form_type if not provided
    if not form_type:
        if rid:
            form_type = 'v2' if is_v2 else 'v1'
        else:
            form_type = '' # Default state
    
    # Always load both V1 and V2 for sidebar display
    configs = ReportConfig.query.all()
    v2_templates = ReportTemplateV2.query.all()
    
    active = None
    active_v2 = None
    excel_html = None
    submissions = []
    
    # --- Permissions & Isolation ---
    role = db.session.get(AppRole, session.get('role_id')) if session.get('role_id') else None
    perms = json.loads(role.perms) if role and role.perms else {}
    is_lead = perms.get('p_stat_lead') or session.get('is_admin')
    is_exec = perms.get('p_stat_exec')
    user_unit = session.get('unit_area')

    # --- Task: Summary Statistics ---
    all_units_query = db.session.query(User.unit_area).distinct()
    if not is_lead:
        all_units_query = all_units_query.filter(User.unit_area == user_unit)
    
    all_units = [u[0] for u in all_units_query.all() if u[0] and u[0] != 'Hệ thống']
    total_units_count = len(all_units)
    reported_unit_set = set()

    if rid:
        if is_v2:
            active_v2 = db.session.get(ReportTemplateV2, rid)
            if active_v2:
                # Latest published version
                ver = (ReportVersionV2.query.filter_by(template_id=rid, is_published=True)
                       .order_by(ReportVersionV2.created_at.desc()).first())
                if ver:
                    raw_subs = (db.session.query(ReportSubmissionV2, User)
                                .join(User, ReportSubmissionV2.user_id == User.id)
                                .filter(ReportSubmissionV2.version_id == ver.id)
                                .all())
                    
                    sub_ids = [s[0].id for s in raw_subs]
                    all_vals = {}
                    if sub_ids:
                        values_query = ReportValueV2.query.filter(ReportValueV2.submission_id.in_(sub_ids)).all()
                        for v in values_query:
                            all_vals[v.cell_key] = v.value

                    for sub, user in raw_subs:
                        u_area = sub.org_unit or user.unit_area
                        reported_unit_set.add(u_area)
                        submissions.append({
                            'id': sub.id,
                            'unit': u_area,
                            'sender': user.fullname,
                            'date': sub.updated_at.strftime('%d/%m/%Y'),
                            'status': sub.status
                        })
                    
                    from excel_renderer import build_v2_stats_table_html
                    metadata = json.loads(ver.metadata_json or '{}')
                    excel_html = build_v2_stats_table_html(ver.excel_file_blob, metadata, all_vals)
        else:
            active = db.session.get(ReportConfig, rid)
            if active:
                # V1 Stats Logic
                try:
                    fields = [f for f in json.loads(active.config_json or '[]')
                              if f.get('is_visible', True)]
                except Exception:
                    fields = []

                raw_query = (db.session.query(ReportData, User)
                       .join(User, ReportData.user_id == User.id)
                       .filter(ReportData.report_id == rid)
                       .order_by(ReportData.report_date.desc(), User.unit_area))

                if not is_lead:
                    raw_query = raw_query.filter(User.unit_area == user_unit)

                raw = raw_query.all()

                seen_units = {}
                for entry, user in raw:
                    try: data = json.loads(entry.data_json or '{}')
                    except Exception: data = {}
                    unit = user.unit_area or user.fullname
                    row = {
                        'unit': unit,
                        'sender': user.fullname,
                        'date': entry.report_date.strftime('%d/%m/%Y') if entry.report_date else '—',
                        'values': {str(f['idx']): data.get(str(f['idx']), '') for f in fields}
                    }
                    if active.is_daily:
                        if unit not in reported_unit_set:
                            reported_unit_set.add(unit)
                            submissions.append(row)
                    else:
                        reported_unit_set.add(unit)
                        submissions.append(row)

                # Render V1 Excel
                from excel_renderer import build_stats_table_html
                excel_html = build_stats_table_html(active.file_blob, active, submissions)

    # Final stats
    sub_count = len(reported_unit_set)
    not_reported_units = [u for u in all_units if u not in reported_unit_set]
    not_reported_count = len(not_reported_units)

    # --- Build Overview Matrix (tổng quan tất cả biểu mẫu × tất cả đơn vị) ---
    # Only computed when no specific form is selected (overview mode)
    overview_matrix = []  # list of { 'form_name', 'version', 'reported': set(), 'not_reported': list() }

    if not rid:
        # V1 overview
        for cfg in configs:
            try:
                reported = set()
                raw_all = (db.session.query(ReportData, User)
                           .join(User, ReportData.user_id == User.id)
                           .filter(ReportData.report_id == cfg.id).all())
                for entry, user in raw_all:
                    u_area = user.unit_area or user.fullname
                    if u_area and u_area != 'Hệ thống':
                        reported.add(u_area)
                not_rep = [u for u in all_units if u not in reported]
                overview_matrix.append({
                    'form_id': cfg.id,
                    'form_name': cfg.name,
                    'version': 'V1',
                    'reported_count': len(reported),
                    'not_reported_count': len(not_rep),
                    'not_reported': not_rep,
                    'url': url_for('forms_bp.stats', rid=cfg.id),
                })
            except Exception:
                pass

        # V2 overview
        for t in v2_templates:
            try:
                reported = set()
                ver = (ReportVersionV2.query.filter_by(template_id=t.id, is_published=True)
                       .order_by(ReportVersionV2.created_at.desc()).first())
                if ver:
                    raw_subs = (db.session.query(ReportSubmissionV2, User)
                                .join(User, ReportSubmissionV2.user_id == User.id)
                                .filter(ReportSubmissionV2.version_id == ver.id).all())
                    for sub, user in raw_subs:
                        u_area = sub.org_unit or user.unit_area
                        if u_area and u_area != 'Hệ thống':
                            reported.add(u_area)
                not_rep = [u for u in all_units if u not in reported]
                overview_matrix.append({
                    'form_id': t.id,
                    'form_name': t.name,
                    'version': 'V2',
                    'reported_count': len(reported),
                    'not_reported_count': len(not_rep),
                    'not_reported': not_rep,
                    'url': url_for('forms_bp.stats', rid=t.id, v2=1),
                })
            except Exception:
                pass

    return render_template('stats_report.html',
                           configs=configs, v2_templates=v2_templates,
                           active=active, active_v2=active_v2,
                           excel_html=excel_html, sub_count=sub_count,
                           submissions=submissions, is_v2=is_v2,
                           total_units_count=total_units_count,
                           not_reported_count=not_reported_count,
                           not_reported_units=not_reported_units,
                           all_units=all_units,
                           form_type=form_type)


@forms_bp.route('/progress')
def progress():
    """Tien do bao cao: bang tong quan tat ca bieu mau x tat ca don vi."""
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    configs = ReportConfig.query.all()
    v2_templates = ReportTemplateV2.query.all()

    # --- Permissions & Isolation ---
    role = db.session.get(AppRole, session.get('role_id')) if session.get('role_id') else None
    perms = json.loads(role.perms) if role and role.perms else {}
    is_lead = perms.get('p_stat_lead') or session.get('is_admin')
    user_unit = session.get('unit_area')

    # Basic stats
    units_query = db.session.query(User.unit_area).distinct()
    if not is_lead:
        units_query = units_query.filter(User.unit_area == user_unit)
    
    all_units = [u[0] for u in units_query.all() if u[0] and u[0] != 'He thong']
    total_units_count = len(all_units)
    overview_matrix = []

    for cfg in configs:
        try:
            reported = set()
            raw_all_query = (db.session.query(ReportData, User)
                       .join(User, ReportData.user_id == User.id)
                       .filter(ReportData.report_id == cfg.id))
            if not is_lead:
                raw_all_query = raw_all_query.filter(User.unit_area == user_unit)
            raw_all = raw_all_query.all()
            for entry, user in raw_all:
                u_area = user.unit_area or user.fullname
                if u_area and u_area not in ('He thong', 'Hệ thống'):
                    reported.add(u_area)
            not_rep = [u for u in all_units if u not in reported]
            overview_matrix.append({
                'form_id': cfg.id, 'form_name': cfg.name, 'version': 'V1',
                'reported_count': len(reported),
                'not_reported_count': len(not_rep), 'not_reported': not_rep,
                'url': url_for('forms_bp.stats', rid=cfg.id),
            })
        except Exception:
            pass

    for t in v2_templates:
        try:
            reported = set()
            ver = (ReportVersionV2.query.filter_by(template_id=t.id, is_published=True)
                   .order_by(ReportVersionV2.created_at.desc()).first())
            if ver:
                raw_subs_query = (db.session.query(ReportSubmissionV2, User)
                            .join(User, ReportSubmissionV2.user_id == User.id)
                            .filter(ReportSubmissionV2.version_id == ver.id))
                if not is_lead:
                    raw_subs_query = raw_subs_query.filter(User.unit_area == user_unit)
                raw_subs = raw_subs_query.all()
                for sub, user in raw_subs:
                    u_area = sub.org_unit or user.unit_area
                    if u_area and u_area not in ('He thong', 'Hệ thống'):
                        reported.add(u_area)
            not_rep = [u for u in all_units if u not in reported]
            overview_matrix.append({
                'form_id': t.id, 'form_name': t.name, 'version': 'V2',
                'reported_count': len(reported),
                'not_reported_count': len(not_rep), 'not_reported': not_rep,
                'url': url_for('forms_bp.stats', rid=t.id, v2=1),
            })
        except Exception:
            pass

    return render_template('progress.html',
                           configs=configs, v2_templates=v2_templates,
                           overview_matrix=overview_matrix,
                           total_units_count=total_units_count,
                           all_units=all_units)


@forms_bp.route('/export-unreported')
def export_unreported():
    """Xuat danh sach chua bao cao (Progress) ra file Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file

    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Don Vi Chua Nop"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="DC2626") # Red color for warning
    
    headers = ["STT", "Biểu mẫu", "Loại", "Danh sách đơn vị chưa nộp"]
    ws.append(headers)
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    configs = ReportConfig.query.all()
    v2_templates = ReportTemplateV2.query.all()

    all_units = [u[0] for u in db.session.query(User.unit_area).distinct().all() if u[0] and u[0] not in ('He thong', 'Hệ thống')]
    
    idx = 1
    for cfg in configs:
        reported = set()
        raw_all = (db.session.query(ReportData, User).join(User, ReportData.user_id == User.id)
                   .filter(ReportData.report_id == cfg.id).all())
        for entry, user in raw_all:
            u_area = user.unit_area or user.fullname
            if u_area and u_area not in ('He thong', 'Hệ thống'): reported.add(u_area)
        not_rep = [u for u in all_units if u not in reported]
        if not_rep:
            ws.append([idx, cfg.name, "V1", "\n".join(not_rep)])
            ws.cell(row=idx+1, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            idx += 1

    for t in v2_templates:
        reported = set()
        ver = (ReportVersionV2.query.filter_by(template_id=t.id, is_published=True)
               .order_by(ReportVersionV2.created_at.desc()).first())
        if ver:
            raw_subs = (db.session.query(ReportSubmissionV2, User).join(User, ReportSubmissionV2.user_id == User.id)
                        .filter(ReportSubmissionV2.version_id == ver.id).all())
            for sub, user in raw_subs:
                u_area = sub.org_unit or user.unit_area
                if u_area and u_area not in ('He thong', 'Hệ thống'): reported.add(u_area)
        not_rep = [u for u in all_units if u not in reported]
        if not_rep:
            ws.append([idx, t.name, "V2", "\n".join(not_rep)])
            ws.cell(row=idx+1, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            idx += 1

    # Adjust widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 80

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"Tien_do_chua_nop_{datetime.now().strftime('%Y%m%d')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@forms_bp.route('/export-form-progress/<string:ftype>/<string:fid>')
def export_form_progress(ftype, fid):
    """Xuat chi tiet tien do cua 1 bieu mau cu the."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file

    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    is_v2 = (ftype.lower() == 'v2')
    form_name = "N/A"
    all_units = [u[0] for u in db.session.query(User.unit_area).distinct().all() if u[0] and u[0] not in ('He thong', 'Hệ thống')]
    
    reported_data = {} # unit -> {name, time}
    
    if is_v2:
        fid_int = int(fid)
        template = db.session.get(ReportTemplateV2, fid_int)
        if not template: return "Not Found", 404
        form_name = template.name
        ver = ReportVersionV2.query.filter_by(template_id=fid_int, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
        if ver:
            subs = db.session.query(ReportSubmissionV2, User).join(User, ReportSubmissionV2.user_id == User.id).filter(ReportSubmissionV2.version_id == ver.id).all()
            for s, u in subs:
                reported_data[s.org_unit or u.unit_area] = {'sender': u.fullname, 'time': s.updated_at.strftime('%H:%M %d/%m/%Y')}
    else:
        cfg = db.session.get(ReportConfig, fid)
        if not cfg: return "Not Found", 404
        form_name = cfg.name
        raw = db.session.query(ReportData, User).join(User, ReportData.user_id == User.id).filter(ReportData.report_id == fid).all()
        for d, u in raw:
            reported_data[u.unit_area] = {'sender': u.fullname, 'time': d.report_date.strftime('%d/%m/%Y') if d.report_date else 'N/A'}

    wb = Workbook()
    ws = wb.active
    ws.title = "Tiến độ chi tiết"
    
    # Headers
    headers = ["STT", "Đơn vị", "Trạng thái", "Người nộp", "Thời gian nộp"]
    ws.append(headers)
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, unit in enumerate(all_units, 1):
        if unit in reported_data:
            ws.append([i, unit, "Đã nộp", reported_data[unit]['sender'], reported_data[unit]['time']])
        else:
            ws.append([i, unit, "Chưa nộp", "-", "-"])
            ws.cell(row=i+1, column=3).font = Font(color="FF0000")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = "".join([c if c.isalnum() else "_" for c in form_name])
    return send_file(output, as_attachment=True, download_name=f"Tien_do_{safe_name}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 80

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    import datetime
    filename = f"Danh_Sach_Don_Vi_Chua_Bao_Cao_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(out, download_name=filename, as_attachment=True)
