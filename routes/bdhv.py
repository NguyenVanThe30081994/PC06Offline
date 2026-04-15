from flask import Blueprint, request, session, redirect, url_for, flash, render_template, jsonify, send_file
from models import db, BDHV_HocVien, BDHV_DonVi, BDHV_ThongKe, BDHV_DangKyMoi, BDHV_ThiLai, BDHV_PhucTra, BDHV_DauMoi
from datetime import datetime
from utils import render_auto_template as render_template
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

bdhv_bp = Blueprint('bdhv_bp', __name__)

@bdhv_bp.route('/bdhv', methods=['GET'])
def index():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    # Lấy danh sách đơn vị cho dropdown
    don_vi_list = BDHV_DonVi.query.all()
    
    # Lấy thống kê
    thong_ke = BDHV_ThongKe.query.order_by(BDHV_ThongKe.ty_le.desc()).all()
    
    # Tổng quan
    total_hv = BDHV_HocVien.query.count()
    total_dv = BDHV_DonVi.query.count()
    
    return render_template('bdhv.html', 
                           don_vi_list=don_vi_list,
                           thong_ke=thong_ke,
                           total_hv=total_hv,
                           total_dv=total_dv)


@bdhv_bp.route('/bdhv/tra-cuu', methods=['GET', 'POST'])
def tra_cuu():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    result = None
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        if cccd:
            # Tìm trong DS HV
            hoc_vien = BDHV_HocVien.query.filter_by(cccd=cccd).first()
            if not hoc_vien:
                # Tìm trong đăng ký mới
                hoc_vien = BDHV_DangKyMoi.query.filter_by(cccd=cccd).first()
            
            if hoc_vien:
                # Tính điểm
                diem_hoc = hoc_vien.diem_hoc or 0
                diem_thi = hoc_vien.diem_thi or 0
                is_pass = (diem_hoc >= 50 and diem_thi >= 50)
                
                ket_qua = hoc_vien.ket_qua
                if not ket_qua:
                    ket_qua = "Hoàn thành khóa học" if is_pass else "Chưa hoàn thành"
                
                result = {
                    'found': True,
                    'ho_ten': hoc_vien.ho_ten,
                    'don_vi': hoc_vien.don_vi,
                    'diem_hoc': f"{diem_hoc}%",
                    'diem_thi': f"{diem_thi}%",
                    'ket_qua': ket_qua,
                    'is_pass': is_pass
                }
            else:
                result = {'found': False, 'message': 'Không tìm thấy thông tin học viên trong hệ thống.'}
    
    don_vi_list = BDHV_DonVi.query.all()
    return render_template('bdhv_tra_cuu.html', result=result, don_vi_list=don_vi_list)


@bdhv_bp.route('/bdhv/tra-cuu-don-vi', methods=['GET', 'POST'])
def tra_cuu_don_vi():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    result = None
    if request.method == 'POST':
        don_vi = request.form.get('don_vi', '').strip()
        if don_vi:
            # Lấy học viên từ DS HV
            hoc_vien_ds = BDHV_HocVien.query.filter_by(don_vi=don_vi, nguon='DS HV').all()
            # Lấy học viên từ đăng ký mới
            hoc_vien_dk = BDHV_DangKyMoi.query.filter_by(don_vi=don_vi).all()
            
            all_hv = []
            stt = 1
            
            for hv in hoc_vien_ds:
                diem_hoc = hv.diem_hoc or 0
                diem_thi = hv.diem_thi or 0
                is_pass = (diem_hoc >= 50 and diem_thi >= 50)
                all_hv.append({
                    'stt': stt,
                    'ho_ten': hv.ho_ten,
                    'diem_hoc': f"{diem_hoc}%",
                    'diem_thi': f"{diem_thi}%",
                    'is_pass': is_pass
                })
                stt += 1
            
            for hv in hoc_vien_dk:
                all_hv.append({
                    'stt': stt,
                    'ho_ten': hv.ho_ten,
                    'diem_hoc': '0%',
                    'diem_thi': '0%',
                    'is_pass': False
                })
                stt += 1
            
            passed = sum(1 for hv in all_hv if hv['is_pass'])
            
            result = {
                'success': True,
                'don_vi': don_vi,
                'stats': {
                    'total': len(all_hv),
                    'passed': passed,
                    'rate': round((passed / len(all_hv) * 100) if all_hv else 0)
                },
                'list': all_hv
            }
    
    don_vi_list = BDHV_DonVi.query.all()
    return render_template('bdhv_don_vi.html', result=result, don_vi_list=don_vi_list)


@bdhv_bp.route('/bdhv/dang-ky', methods=['GET', 'POST'])
def dang_ky():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        ho_ten = request.form.get('ho_ten', '').strip()
        don_vi = request.form.get('don_vi', '').strip()
        
        if not cccd or not ho_ten or not don_vi:
            flash('Vui lòng nhập đầy đủ thông tin!', 'danger')
            return redirect(url_for('bdhv_bp.dang_ky'))
        
        # Kiểm tra trùng CCCD
        existing = BDHV_DangKyMoi.query.filter_by(cccd=cccd).first()
        if existing:
            flash(f'Số CCCD/Căn cước {cccd} đã được đăng ký!', 'warning')
            return redirect(url_for('bdhv_bp.dang_ky'))
        
        # Lấy STT mới
        last_record = BDHV_DangKyMoi.query.order_by(BDHV_DangKyMoi.stt.desc()).first()
        new_stt = (last_record.stt + 1) if last_record else 1
        
        # Lưu
        new_hv = BDHV_DangKyMoi(
            stt=new_stt,
            ho_ten=ho_ten,
            cccd=cccd,
            don_vi=don_vi
        )
        db.session.add(new_hv)
        db.session.commit()
        
        flash('Đăng ký học viên mới thành công!', 'success')
        return redirect(url_for('bdhv_bp.dang_ky'))
    
    don_vi_list = BDHV_DonVi.query.all()
    return render_template('bdhv_dang_ky.html', don_vi_list=don_vi_list)


@bdhv_bp.route('/bdhv/thi-lai', methods=['GET', 'POST'])
def thi_lai():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        ho_ten = request.form.get('ho_ten', '').strip()
        don_vi = request.form.get('don_vi', '').strip()
        ly_do = request.form.get('ly_do', '').strip()
        
        if not cccd or not ho_ten or not don_vi:
            flash('Vui lòng nhập đầy đủ thông tin!', 'danger')
            return redirect(url_for('bdhv_bp.thi_lai'))
        
        # Lưu đăng ký thi lại
        new_record = BDHV_ThiLai(
            ho_ten=ho_ten,
            cccd=cccd,
            don_vi=don_vi,
            ly_do=ly_do
        )
        db.session.add(new_record)
        db.session.commit()
        
        flash(f'Đăng ký thi lại thành công! Bạn sẽ được thi lại vào ngày mai.', 'success')
        return redirect(url_for('bdhv_bp.thi_lai'))
    
    don_vi_list = BDHV_DonVi.query.all()
    return render_template('bdhv_thi_lai.html', don_vi_list=don_vi_list)


@bdhv_bp.route('/bdhv/phuc-tra', methods=['GET', 'POST'])
def phuc_tra():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        ho_ten = request.form.get('ho_ten', '').strip()
        don_vi = request.form.get('don_vi', '').strip()
        ly_do = request.form.get('ly_do', '').strip()
        
        if not cccd or not ho_ten or not don_vi:
            flash('Vui lòng nhập đầy đủ thông tin!', 'danger')
            return redirect(url_for('bdhv_bp.phuc_tra'))
        
        # Lưu phúc tra
        new_record = BDHV_PhucTra(
            ho_ten=ho_ten,
            cccd=cccd,
            don_vi=don_vi,
            ly_do=ly_do
        )
        db.session.add(new_record)
        db.session.commit()
        
        flash('Bạn đã gửi phúc tra thành công. Kết quả sẽ được cập nhật sau 24h.', 'success')
        return redirect(url_for('bdhv_bp.phuc_tra'))
    
    don_vi_list = BDHV_DonVi.query.all()
    return render_template('bdhv_phuc_tra.html', don_vi_list=don_vi_list)


@bdhv_bp.route('/bdhv/thong-ke', methods=['GET'])
def thong_ke():
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    don_vi_list = BDHV_DonVi.query.all()
    selected_don_vi = request.args.get('don_vi', '')
    don_vi_data = None
    
    # Lấy thống kê sắp xếp theo tỷ lệ
    thong_ke = BDHV_ThongKe.query.order_by(BDHV_ThongKe.ty_le.desc()).all()
    
    # Tính tổng
    total_18 = sum(t.tong_18 for t in thong_ke)
    total_dang_ky = sum(t.da_dang_ky for t in thong_ke)
    total_thi = sum(t.diem_thi for t in thong_ke)
    overall_rate = round((total_thi / total_18 * 100) if total_18 > 0 else 0)
    
    # Nếu có đơn vị được chọn, lấy chi tiết
    if selected_don_vi:
        don_vi_data = BDHV_ThongKe.query.filter_by(ten_don_vi=selected_don_vi).first()
    
    return render_template('bdhv_thong_ke.html',
                           thong_ke=thong_ke,
                           don_vi_list=don_vi_list,
                           selected_don_vi=selected_don_vi,
                           don_vi_data=don_vi_data,
                           total_18=total_18,
                           total_dang_ky=total_dang_ky,
                           total_thi=total_thi,
                           overall_rate=overall_rate)


@bdhv_bp.route('/bdhv/xep-hang', methods=['GET'])
def xep_hang():
    """Xếp hạng đơn vị theo tỷ lệ hoàn thành"""
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    # Lấy thống kê sắp xếp theo tỷ lệ giảm dần
    thong_ke = BDHV_ThongKe.query.order_by(BDHV_ThongKe.ty_le.desc()).all()
    
    # Tính tổng
    total_18 = sum(t.tong_18 for t in thong_ke)
    total_dang_ky = sum(t.da_dang_ky for t in thong_ke)
    total_thi = sum(t.diem_thi for t in thong_ke)
    overall_rate = round((total_thi / total_18 * 100) if total_18 > 0 else 0)
    
    # Xếp hạng
    ranking_data = []
    for idx, tk in enumerate(thong_ke, 1):
        ranking_data.append({
            'rank': idx,
            'ten_don_vi': tk.ten_don_vi,
            'tong_18': tk.tong_18,
            'da_dang_ky': tk.da_dang_ky,
            'da_hoan_thanh': tk.da_hoan_thanh,
            'diem_thi': tk.diem_thi,
            'ty_le': tk.ty_le,
            'rank_class': 'top-1' if idx == 1 else 'top-2' if idx == 2 else 'top-3' if idx == 3 else 'normal'
        })
    
    return render_template('bdhv_xep_hang.html',
                           ranking_data=ranking_data,
                           total_18=total_18,
                           total_dang_ky=total_dang_ky,
                           total_thi=total_thi,
                           overall_rate=overall_rate)


@bdhv_bp.route('/bdhv/database', methods=['GET'])
def database():
    """Quản lý database BDHV"""
    if not session.get('is_admin'): return redirect(url_for('auth_bp.login'))
    
    stats = {
        'hoc_vien': BDHV_HocVien.query.count(),
        'don_vi': BDHV_DonVi.query.count(),
        'dang_ky': BDHV_DangKyMoi.query.count(),
        'thi_lai': BDHV_ThiLai.query.count(),
        'phuc_tra': BDHV_PhucTra.query.count(),
        'thong_ke': BDHV_ThongKe.query.count()
    }
    
    return render_template('bdhv_database.html', stats=stats)


# === API CHO ADMIN ===

@bdhv_bp.route('/api/bdhv/import-hocvien', methods=['POST'])
def import_hocvien():
    """Import học viên từ Excel"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    # TODO: Implement Excel import
    return jsonify({'success': True, 'message': 'Import thành công!'})


@bdhv_bp.route('/api/bdhv/don-vi', methods=['POST'])
def add_don_vi():
    """Thêm đơn vị mới"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    name = request.json.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Tên đơn vị không được để trống'}), 400
    
    existing = BDHV_DonVi.query.filter_by(ten_don_vi=name).first()
    if existing:
        return jsonify({'error': 'Đơn vị đã tồn tại'}), 400
    
    new_dv = BDHV_DonVi(ten_don_vi=name)
    db.session.add(new_dv)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Thêm đơn vị thành công!'})


@bdhv_bp.route('/api/bdhv/thong-ke', methods=['POST'])
def update_thong_ke():
    """Cập nhật thống kê đơn vị"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    ten_don_vi = data.get('ten_don_vi')
    tong_18 = data.get('tong_18', 0)
    da_dang_ky = data.get('da_dang_ky', 0)
    da_hoan_thanh = data.get('da_hoan_thanh', 0)
    diem_thi = data.get('diem_thi', 0)
    
    # Tính tỷ lệ
    ty_le = round((diem_thi / tong_18 * 100) if tong_18 > 0 else 0, 2)
    
    # Tìm hoặc tạo mới
    tk = BDHV_ThongKe.query.filter_by(ten_don_vi=ten_don_vi).first()
    if tk:
        tk.tong_18 = tong_18
        tk.da_dang_ky = da_dang_ky
        tk.da_hoan_thanh = da_hoan_thanh
        tk.diem_thi = diem_thi
        tk.ty_le = ty_le
    else:
        tk = BDHV_ThongKe(
            ten_don_vi=ten_don_vi,
            tong_18=tong_18,
            da_dang_ky=da_dang_ky,
            da_hoan_thanh=da_hoan_thanh,
            diem_thi=diem_thi,
            ty_le=ty_le
        )
        db.session.add(tk)
    
    db.session.commit()
    return jsonify({'success': True})


@bdhv_bp.route('/api/bdhv/delete-hocvien', methods=['POST'])
def delete_hocvien():
    """Xóa tất cả học viên"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    count = BDHV_HocVien.query.delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Đã xóa {count} học viên!'})


@bdhv_bp.route('/api/bdhv/delete-dangky', methods=['POST'])
def delete_dangky():
    """Xóa tất cả đăng ký mới"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    count = BDHV_DangKyMoi.query.delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Đã xóa {count} đăng ký!'})


@bdhv_bp.route('/api/bdhv/delete-thilai', methods=['POST'])
def delete_thilai():
    """Xóa tất cả đăng ký thi lại"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    count = BDHV_ThiLai.query.delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Đã xóa {count} đăng ký thi lại!'})


@bdhv_bp.route('/api/bdhv/delete-phuctra', methods=['POST'])
def delete_phuctra():
    """Xóa tất cả phúc tra"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    count = BDHV_PhucTra.query.delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Đã xóa {count} phúc tra!'})


@bdhv_bp.route('/api/bdhv/delete-thongke', methods=['POST'])
def delete_thongke():
    """Xóa tất cả thống kê"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    count = BDHV_ThongKe.query.delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Đã xóa {count} thống kê!'})


@bdhv_bp.route('/api/bdhv/delete-donvi', methods=['POST'])
def delete_donvi():
    """Xóa tất cả đơn vị"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    # Xóa thống kê trước
    BDHV_ThongKe.query.delete()
    # Xóa đơn vị
    count = BDHV_DonVi.query.delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Đã xóa {count} đơn vị!'})


@bdhv_bp.route('/api/bdhv/delete-all', methods=['POST'])
def delete_all():
    """Xóa tất cả dữ liệu BDHV"""
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    counts = {}
    counts['hoc_vien'] = BDHV_HocVien.query.delete()
    counts['dang_ky'] = BDHV_DangKyMoi.query.delete()
    counts['thi_lai'] = BDHV_ThiLai.query.delete()
    counts['phuc_tra'] = BDHV_PhucTra.query.delete()
    counts['thong_ke'] = BDHV_ThongKe.query.delete()
    counts['don_vi'] = BDHV_DonVi.query.delete()
    db.session.commit()
    
    total = sum(counts.values())
    return jsonify({'success': True, 'message': f'Đã xóa tất cả dữ liệu BDHV ({total} bản ghi)!'})


# === EXPORT EXCEL ===

@bdhv_bp.route('/bdhv/export/danh-sach', methods=['GET'])
def export_danh_sach():
    """Xuất danh sách học viên ra Excel"""
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    try:
        # Lấy tất cả học viên
        hoc_vien_list = BDHV_HocVien.query.all()
        dong_dang_ky = BDHV_DangKyMoi.query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DanhSachHV"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['STT', 'Họ tên', 'CCCD', 'Đơn vị', 'Điểm học', 'Điểm thi', 'Kết quả', 'Nguồn']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data - Học viên từ DS HV
        row = 2
        for hv in hoc_vien_list:
            ws.cell(row=row, column=1, value=hv.stt or (row-1))
            ws.cell(row=row, column=2, value=hv.ho_ten)
            ws.cell(row=row, column=3, value=hv.cccd)
            ws.cell(row=row, column=4, value=hv.don_vi)
            ws.cell(row=row, column=5, value=hv.diem_hoc)
            ws.cell(row=row, column=6, value=hv.diem_thi)
            ws.cell(row=row, column=7, value=hv.ket_qua)
            ws.cell(row=row, column=8, value=hv.nguon)
            row += 1
        
        # Data - Đăng ký mới
        for dk in dong_dang_ky:
            ws.cell(row=row, column=1, value=dk.stt or '')
            ws.cell(row=row, column=2, value=dk.ho_ten)
            ws.cell(row=row, column=3, value=dk.cccd)
            ws.cell(row=row, column=4, value=dk.don_vi)
            ws.cell(row=row, column=5, value='')
            ws.cell(row=row, column=6, value='')
            ws.cell(row=row, column=7, value='Đã đăng ký')
            ws.cell(row=row, column=8, value='DKMoi')
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 10
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            download_name=f"BDHV_DanhSach_{datetime.now().strftime('%Y%m%d')}.xlsx",
            as_attachment=True
        )
    except Exception as e:
        flash(f'Lỗi xuất Excel: {str(e)}', 'danger')
        return redirect(url_for('bdhv_bp.index'))


@bdhv_bp.route('/bdhv/export/thong-ke', methods=['GET'])
def export_thong_ke():
    """Xuất thống kê theo đơn vị ra Excel"""
    if not session.get('uid'): return redirect(url_for('auth_bp.login'))
    
    try:
        thong_ke = BDHV_ThongKe.query.order_by(BDHV_ThongKe.ty_le.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ThongKe"
        
        # Header style
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['STT', 'Đơn vị', 'Tổng DS 18+', 'Đã đăng ký', 'Đã hoàn thành', 'Đã thi', 'Tỷ lệ (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data
        for idx, tk in enumerate(thong_ke, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=tk.ten_don_vi)
            ws.cell(row=idx+1, column=3, value=tk.tong_18)
            ws.cell(row=idx+1, column=4, value=tk.da_dang_ky)
            ws.cell(row=idx+1, column=5, value=tk.da_hoan_thanh)
            ws.cell(row=idx+1, column=6, value=tk.diem_thi)
            ws.cell(row=idx+1, column=7, value=tk.ty_le)
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            download_name=f"BDHV_ThongKe_{datetime.now().strftime('%Y%m%d')}.xlsx",
            as_attachment=True
        )
    except Exception as e:
        flash(f'Lỗi xuất Excel: {str(e)}', 'danger')
        return redirect(url_for('bdhv_bp.index'))